from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import shutil
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Email Configuration
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.office365.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER)

# File upload configuration
UPLOAD_DIR = Path('/app/uploads')
UPLOAD_DIR.mkdir(exist_ok=True)

security = HTTPBearer()

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ===== MODELS =====

class UserRole:
    CIUDADANO = "ciudadano"
    ATENCION_USUARIO = "atencion_usuario"
    GESTOR = "gestor"
    GESTOR_AUXILIAR = "gestor_auxiliar"
    COORDINADOR = "coordinador"
    ADMINISTRADOR = "administrador"

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    full_name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    full_name: str
    role: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserRoleUpdate(BaseModel):
    user_id: str
    new_role: str

class PetitionStatus:
    RADICADO = "radicado"
    ASIGNADO = "asignado"
    RECHAZADO = "rechazado"
    REVISION = "revision"
    DEVUELTO = "devuelto"
    FINALIZADO = "finalizado"

class PetitionCreate(BaseModel):
    nombre_completo: str
    correo: EmailStr
    telefono: str
    tipo_tramite: str
    municipio: str

class PetitionUpdate(BaseModel):
    nombre_completo: Optional[str] = None
    correo: Optional[EmailStr] = None
    telefono: Optional[str] = None
    tipo_tramite: Optional[str] = None
    municipio: Optional[str] = None
    estado: Optional[str] = None
    notas: Optional[str] = None
    gestor_id: Optional[str] = None

class GestorAssignment(BaseModel):
    petition_id: str
    gestor_id: str
    is_auxiliar: bool = False

class Petition(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    radicado: str
    user_id: str
    nombre_completo: str
    correo: str
    telefono: str
    tipo_tramite: str
    municipio: str
    estado: str = PetitionStatus.RADICADO
    notas: str = ""
    gestor_id: Optional[str] = None
    gestores_asignados: List[str] = []
    archivos: List[dict] = []
    historial: List[dict] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# ===== PREDIO MODELS (Codigo Nacional Catastral) =====

# Catalogo de municipios con codigo catastral nacional
MUNICIPIOS_DIVIPOLA = {
    "Ábrego": {"departamento": "54", "municipio": "003"},
    "Bucarasica": {"departamento": "54", "municipio": "109"},
    "Cáchira": {"departamento": "54", "municipio": "128"},
    "Convención": {"departamento": "54", "municipio": "206"},
    "El Carmen": {"departamento": "54", "municipio": "245"},
    "El Tarra": {"departamento": "54", "municipio": "250"},
    "Hacarí": {"departamento": "54", "municipio": "344"},
    "La Playa": {"departamento": "54", "municipio": "398"},
    "Río de Oro": {"departamento": "47", "municipio": "545"},  # Cesar
    "San Calixto": {"departamento": "54", "municipio": "670"},
    "Sardinata": {"departamento": "54", "municipio": "720"},
    "Teorama": {"departamento": "54", "municipio": "800"}
}

# Catálogo de destino económico
DESTINO_ECONOMICO = {
    "A": "Habitacional",
    "B": "Industrial",
    "C": "Comercial",
    "D": "Agropecuario",
    "E": "Minero",
    "F": "Recreacional",
    "G": "Salubridad",
    "H": "Institucional",
    "I": "Educativo",
    "J": "Religioso",
    "K": "Cultural",
    "L": "Lote",
    "M": "Pecuario",
    "N": "Agrícola",
    "O": "Uso Público",
    "P": "Forestal",
    "Q": "Mixto Comercial-Habitacional",
    "R": "Servicios Especiales",
    "S": "Institucional Público",
    "0": "Sin clasificar"
}

# Catálogo de tipo de documento
TIPO_DOCUMENTO_PREDIO = {
    "C": "Cédula de Ciudadanía",
    "E": "Cédula de Extranjería",
    "N": "NIT",
    "T": "Tarjeta de Identidad",
    "P": "Pasaporte",
    "X": "Sin documento / Entidad"
}

# Catálogo de estado civil
ESTADO_CIVIL_PREDIO = {
    "S": "Soltero/a",
    "E": "Casado/a con sociedad conyugal",
    "D": "Casado/a sin sociedad conyugal",
    "V": "Separación de bienes",
    "U": "Unión marital de hecho"
}

class PredioR1Create(BaseModel):
    """Registro R1 - Información Jurídica del Predio"""
    municipio: str
    zona: str = "00"  # 00=Rural, 01+=Urbano
    sector: str = "01"
    manzana_vereda: str = "0000"
    terreno: str = "0001"
    condicion_predio: str = "0000"
    predio_horizontal: str = "0000"
    
    # Propietario
    nombre_propietario: str
    tipo_documento: str
    numero_documento: str
    estado_civil: Optional[str] = None
    
    # Ubicación y características
    direccion: str
    comuna: str = "0"
    destino_economico: str
    area_terreno: float
    area_construida: float = 0
    avaluo: float
    
    # Mutación
    tipo_mutacion: Optional[str] = None
    numero_resolucion: Optional[str] = None
    fecha_resolucion: Optional[str] = None

class PredioR2Create(BaseModel):
    """Registro R2 - Información Física del Predio"""
    matricula_inmobiliaria: Optional[str] = None
    
    # Zona 1
    zona_fisica_1: float = 0
    zona_economica_1: float = 0
    area_terreno_1: float = 0
    
    # Zona 2
    zona_fisica_2: float = 0
    zona_economica_2: float = 0
    area_terreno_2: float = 0
    
    # Construcción 1
    habitaciones_1: int = 0
    banos_1: int = 0
    locales_1: int = 0
    pisos_1: int = 1
    tipificacion_1: float = 0
    uso_1: int = 0
    puntaje_1: float = 0
    area_construida_1: float = 0
    
    # Construcción 2
    habitaciones_2: int = 0
    banos_2: int = 0
    locales_2: int = 0
    pisos_2: int = 0
    tipificacion_2: float = 0
    uso_2: int = 0
    puntaje_2: float = 0
    area_construida_2: float = 0

class PredioCreate(BaseModel):
    r1: PredioR1Create
    r2: Optional[PredioR2Create] = None

class PredioUpdate(BaseModel):
    # R1 fields
    nombre_propietario: Optional[str] = None
    tipo_documento: Optional[str] = None
    numero_documento: Optional[str] = None
    estado_civil: Optional[str] = None
    direccion: Optional[str] = None
    comuna: Optional[str] = None
    destino_economico: Optional[str] = None
    area_terreno: Optional[float] = None
    area_construida: Optional[float] = None
    avaluo: Optional[float] = None
    tipo_mutacion: Optional[str] = None
    numero_resolucion: Optional[str] = None
    fecha_resolucion: Optional[str] = None
    
    # R2 fields
    matricula_inmobiliaria: Optional[str] = None
    zona_fisica_1: Optional[float] = None
    zona_economica_1: Optional[float] = None
    area_terreno_1: Optional[float] = None
    habitaciones_1: Optional[int] = None
    banos_1: Optional[int] = None
    locales_1: Optional[int] = None
    pisos_1: Optional[int] = None
    puntaje_1: Optional[float] = None
    area_construida_1: Optional[float] = None


# ===== SISTEMA DE APROBACIÓN DE PREDIOS =====

class PredioEstadoAprobacion:
    """Estados de aprobación para cambios en predios"""
    APROBADO = "aprobado"  # Cambios aplicados y firmes
    PENDIENTE_CREACION = "pendiente_creacion"  # Nuevo predio esperando aprobación
    PENDIENTE_MODIFICACION = "pendiente_modificacion"  # Modificación esperando aprobación
    PENDIENTE_ELIMINACION = "pendiente_eliminacion"  # Eliminación esperando aprobación
    RECHAZADO = "rechazado"  # Cambio rechazado por coordinador

class CambioPendienteCreate(BaseModel):
    """Modelo para crear un cambio pendiente"""
    predio_id: Optional[str] = None  # None para nuevos predios
    tipo_cambio: str  # creacion, modificacion, eliminacion
    datos_propuestos: dict  # Datos del predio (nuevo o modificado)
    justificacion: Optional[str] = None

class CambioAprobacionRequest(BaseModel):
    """Modelo para aprobar/rechazar un cambio"""
    cambio_id: str
    aprobado: bool
    comentario: Optional[str] = None


# ===== UTILITY FUNCTIONS =====

import re

def validate_password(password: str) -> tuple[bool, str]:
    """
    Validate password requirements:
    - Minimum 6 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one digit
    - Special characters are allowed: !@#$%^&*()_+-=[]{}|;':\",./<>?
    Returns: (is_valid, error_message)
    """
    if len(password) < 6:
        return False, "La contraseña debe tener al menos 6 caracteres"
    
    if not re.search(r'[A-Z]', password):
        return False, "La contraseña debe contener al menos una letra mayúscula"
    
    if not re.search(r'[a-z]', password):
        return False, "La contraseña debe contener al menos una letra minúscula"
    
    if not re.search(r'\d', password):
        return False, "La contraseña debe contener al menos un número"
    
    # Allow special characters - password is valid if it passes above checks
    return True, ""

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token inválido")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuario no encontrado")
    return user

async def generate_radicado() -> str:
    now = datetime.now()
    date_str = now.strftime("%d-%m-%Y")
    
    # Get count for today
    start_of_day = datetime(now.year, now.month, now.day, tzinfo=timezone.utc)
    count = await db.petitions.count_documents({
        "created_at": {"$gte": start_of_day.isoformat()}
    })
    
    sequence = str(count + 1).zfill(4)
    return f"RASMCG-{sequence}-{date_str}"

async def send_email(to_email: str, subject: str, body: str):
    if not SMTP_USER or not SMTP_PASSWORD:
        logging.warning("SMTP credentials not configured, skipping email")
        return
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        logging.info(f"Email sent to {to_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {str(e)}")


# ===== AUTH ROUTES =====

@api_router.post("/auth/register")
async def register(user_data: UserRegister):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El correo ya está registrado")
    
    # Validate password
    is_valid, error_msg = validate_password(user_data.password)
    if not is_valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_msg)
    
    # Always assign ciudadano role on self-registration
    user = User(
        email=user_data.email,
        full_name=user_data.full_name,
        role=UserRole.CIUDADANO
    )
    
    doc = user.model_dump()
    doc['password'] = hash_password(user_data.password)
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    token = create_token(user.id, user.email, user.role)
    
    return {
        "token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "full_name": user.full_name,
            "role": user.role
        }
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales inválidas")
    
    if not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales inválidas")
    
    token = create_token(user['id'], user['email'], user['role'])
    
    return {
        "token": token,
        "user": {
            "id": user['id'],
            "email": user['email'],
            "full_name": user['full_name'],
            "role": user['role']
        }
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    # Get additional user data from database
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    return {
        "id": current_user['id'],
        "email": current_user['email'],
        "full_name": current_user['full_name'],
        "role": current_user['role'],
        "puede_actualizar_gdb": user_db.get('puede_actualizar_gdb', False) if user_db else False
    }


# ===== PASSWORD RECOVERY =====

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Send password reset email"""
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No existe una cuenta con ese correo")
    
    # Check if SMTP is configured
    if not SMTP_USER or not SMTP_PASSWORD:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="El servicio de correo no está configurado")
    
    # Generate reset token (valid for 1 hour)
    reset_token = str(uuid.uuid4())
    expiration = datetime.now(timezone.utc) + timedelta(hours=1)
    
    # Store reset token in database
    await db.password_resets.delete_many({"email": request.email})  # Remove old tokens
    await db.password_resets.insert_one({
        "email": request.email,
        "token": reset_token,
        "expires_at": expiration.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Get frontend URL from environment or use default
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"
    
    # Send email
    email_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background-color: #047857; padding: 20px; border-radius: 8px 8px 0 0;">
            <h1 style="color: white; margin: 0;">Asomunicipios</h1>
            <p style="color: #d1fae5; margin: 5px 0 0 0;">Sistema de Gestión Catastral</p>
        </div>
        <div style="background-color: #f8fafc; padding: 30px; border-radius: 0 0 8px 8px;">
            <h2 style="color: #1e293b; margin-top: 0;">Recuperación de Contraseña</h2>
            <p style="color: #475569;">Hola <strong>{user['full_name']}</strong>,</p>
            <p style="color: #475569;">Hemos recibido una solicitud para restablecer la contraseña de tu cuenta.</p>
            <p style="color: #475569;">Haz clic en el siguiente botón para crear una nueva contraseña:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_link}" style="background-color: #047857; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold;">Restablecer Contraseña</a>
            </div>
            <p style="color: #64748b; font-size: 14px;">Este enlace expirará en 1 hora.</p>
            <p style="color: #64748b; font-size: 14px;">Si no solicitaste este cambio, puedes ignorar este correo.</p>
            <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 20px 0;">
            <p style="color: #94a3b8; font-size: 12px; margin: 0;">Este es un mensaje automático, por favor no responda a este correo.</p>
        </div>
    </body>
    </html>
    """
    
    try:
        await send_email(request.email, "Recuperación de Contraseña - Asomunicipios", email_body)
    except Exception as e:
        logging.error(f"Failed to send reset email: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al enviar el correo")
    
    return {"message": "Se ha enviado un enlace de recuperación a tu correo"}

@api_router.get("/auth/validate-reset-token")
async def validate_reset_token(token: str):
    """Validate password reset token"""
    reset_record = await db.password_resets.find_one({"token": token}, {"_id": 0})
    
    if not reset_record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Token inválido")
    
    expires_at = datetime.fromisoformat(reset_record['expires_at'])
    if datetime.now(timezone.utc) > expires_at:
        await db.password_resets.delete_one({"token": token})
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El token ha expirado")
    
    return {"valid": True}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """Reset password with token"""
    reset_record = await db.password_resets.find_one({"token": request.token}, {"_id": 0})
    
    if not reset_record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Token inválido")
    
    expires_at = datetime.fromisoformat(reset_record['expires_at'])
    if datetime.now(timezone.utc) > expires_at:
        await db.password_resets.delete_one({"token": request.token})
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El token ha expirado")
    
    # Validate new password
    is_valid, error_msg = validate_password(request.new_password)
    if not is_valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_msg)
    
    # Update password
    new_hashed_password = hash_password(request.new_password)
    await db.users.update_one(
        {"email": reset_record['email']},
        {"$set": {"password": new_hashed_password}}
    )
    
    # Delete used token
    await db.password_resets.delete_one({"token": request.token})
    
    return {"message": "Contraseña actualizada exitosamente"}


# ===== USER MANAGEMENT ROUTES =====

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: dict = Depends(get_current_user)):
    # Only admin, coordinador, and atencion can view users
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return users

@api_router.patch("/users/role")
async def update_user_role(role_update: UserRoleUpdate, current_user: dict = Depends(get_current_user)):
    # Only admin, coordinador, and atencion can change roles
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para cambiar roles")
    
    # Validate new role
    valid_roles = [UserRole.CIUDADANO, UserRole.ATENCION_USUARIO, UserRole.GESTOR, UserRole.GESTOR_AUXILIAR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]
    if role_update.new_role not in valid_roles:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rol inválido")
    
    user = await db.users.find_one({"id": role_update.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    
    await db.users.update_one(
        {"id": role_update.user_id},
        {"$set": {"role": role_update.new_role}}
    )
    
    return {"message": "Rol actualizado exitosamente", "new_role": role_update.new_role}


# ===== PETITION ROUTES =====

@api_router.post("/petitions")
async def create_petition(
    nombre_completo: str = Form(...),
    correo: str = Form(...),
    telefono: str = Form(...),
    tipo_tramite: str = Form(...),
    municipio: str = Form(...),
    files: List[UploadFile] = File(default=[]),
    current_user: dict = Depends(get_current_user)
):
    radicado = await generate_radicado()
    
    # Save files
    saved_files = []
    for file in files:
        if file.filename:
            file_id = str(uuid.uuid4())
            file_ext = Path(file.filename).suffix
            file_path = UPLOAD_DIR / f"{file_id}{file_ext}"
            
            with file_path.open("wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            saved_files.append({
                "id": file_id,
                "original_name": file.filename,
                "path": str(file_path)
            })
    
    # Initialize historial
    historial = [{
        "accion": "Radicado creado",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": None,
        "estado_nuevo": PetitionStatus.RADICADO,
        "notas": "Petición radicada en el sistema",
        "fecha": datetime.now(timezone.utc).isoformat()
    }]
    
    petition = Petition(
        radicado=radicado,
        user_id=current_user['id'],
        nombre_completo=nombre_completo,
        correo=correo,
        telefono=telefono,
        tipo_tramite=tipo_tramite,
        municipio=municipio,
        archivos=saved_files,
        historial=historial
    )
    
    doc = petition.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.petitions.insert_one(doc)
    
    # Send email notification to atencion usuarios only if created by citizen
    if current_user['role'] == UserRole.CIUDADANO:
        atencion_users = await db.users.find({"role": UserRole.ATENCION_USUARIO}, {"_id": 0}).to_list(100)
        for user in atencion_users:
            await send_email(
                user['email'],
                f"Nueva Petición - {radicado}",
                f"<h3>Nueva petición radicada</h3><p>Radicado: {radicado}</p><p>Solicitante: {nombre_completo}</p><p>Tipo: {tipo_tramite}</p>"
            )
    
    return petition

@api_router.get("/petitions")
async def get_petitions(current_user: dict = Depends(get_current_user)):
    # Citizens only see their own petitions
    if current_user['role'] == UserRole.CIUDADANO:
        query = {"user_id": current_user['id']}
    # Gestores see assigned petitions
    elif current_user['role'] in [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]:
        query = {"gestores_asignados": current_user['id']}
    else:
        # Staff can see all petitions
        query = {}
    
    petitions = await db.petitions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for petition in petitions:
        if isinstance(petition['created_at'], str):
            petition['created_at'] = datetime.fromisoformat(petition['created_at'])
        if isinstance(petition['updated_at'], str):
            petition['updated_at'] = datetime.fromisoformat(petition['updated_at'])
    
    return petitions

@api_router.get("/petitions/{petition_id}")
async def get_petition(petition_id: str, current_user: dict = Depends(get_current_user)):
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Citizens can only see their own petitions
    if current_user['role'] == UserRole.CIUDADANO and petition['user_id'] != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para ver esta petición")
    
    # Gestores can only see assigned petitions
    if current_user['role'] in [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]:
        if current_user['id'] not in petition.get('gestores_asignados', []):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para ver esta petición")
    
    if isinstance(petition['created_at'], str):
        petition['created_at'] = datetime.fromisoformat(petition['created_at'])
    if isinstance(petition['updated_at'], str):
        petition['updated_at'] = datetime.fromisoformat(petition['updated_at'])
    
    return petition

@api_router.post("/petitions/{petition_id}/upload")
async def upload_petition_files(
    petition_id: str,
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user)
):
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Citizens and staff can upload files
    # Citizens: only to their own petitions
    # Staff: to any petition they have access to
    if current_user['role'] == UserRole.CIUDADANO and petition['user_id'] != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    saved_files = []
    for file in files:
        if file.filename:
            file_id = str(uuid.uuid4())
            file_ext = Path(file.filename).suffix
            file_path = UPLOAD_DIR / f"{file_id}{file_ext}"
            
            with file_path.open("wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            saved_files.append({
                "id": file_id,
                "original_name": file.filename,
                "path": str(file_path),
                "uploaded_by": current_user['id'],
                "uploaded_by_name": current_user['full_name'],
                "uploaded_by_role": current_user['role'],
                "upload_date": datetime.now(timezone.utc).isoformat()
            })
    
    current_files = petition.get('archivos', [])
    updated_files = current_files + saved_files
    
    # Add to historial
    uploader_role = "Ciudadano" if current_user['role'] == UserRole.CIUDADANO else current_user['role'].replace('_', ' ').title()
    historial_entry = {
        "accion": f"Archivos cargados por {uploader_role} ({len(saved_files)} archivo(s))",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": petition['estado'],
        "estado_nuevo": petition['estado'],
        "notas": f"Se cargaron {len(saved_files)} archivo(s) adicional(es)",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    
    await db.petitions.update_one(
        {"id": petition_id},
        {"$set": {
            "archivos": updated_files,
            "historial": current_historial,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Notify based on who uploaded
    if current_user['role'] == UserRole.CIUDADANO:
        # Notify assigned gestores or atencion usuario
        if petition.get('gestores_asignados'):
            for gestor_id in petition['gestores_asignados']:
                gestor = await db.users.find_one({"id": gestor_id}, {"_id": 0})
                if gestor:
                    await send_email(
                        gestor['email'],
                        f"Nuevos archivos - {petition['radicado']}",
                        f"<h3>El ciudadano ha cargado nuevos archivos</h3><p>Radicado: {petition['radicado']}</p>"
                    )
        else:
            atencion_users = await db.users.find({"role": UserRole.ATENCION_USUARIO}, {"_id": 0}).to_list(100)
            for user in atencion_users:
                await send_email(
                    user['email'],
                    f"Nuevos archivos - {petition['radicado']}",
                    f"<h3>El ciudadano ha cargado nuevos archivos</h3><p>Radicado: {petition['radicado']}</p>"
                )
    else:
        # Notify citizen if staff uploaded
        citizen = await db.users.find_one({"id": petition['user_id']}, {"_id": 0})
        if citizen:
            await send_email(
                citizen['email'],
                f"Nuevos documentos disponibles - {petition['radicado']}",
                f"<h3>Se han agregado nuevos documentos a su trámite</h3><p>Radicado: {petition['radicado']}</p><p>Puede descargarlos desde el sistema.</p>"
            )
    
    return {"message": "Archivos subidos exitosamente", "files": saved_files}


@api_router.get("/petitions/{petition_id}/download-zip")
async def download_citizen_files_as_zip(petition_id: str, current_user: dict = Depends(get_current_user)):
    """Download all files uploaded by citizen as a ZIP file"""
    # Only staff can download citizen files
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Filter files uploaded by citizen
    citizen_files = []
    for archivo in petition.get('archivos', []):
        uploaded_by_role = archivo.get('uploaded_by_role', 'ciudadano')
        if uploaded_by_role == UserRole.CIUDADANO or not archivo.get('uploaded_by_role'):
            # If no uploaded_by_role, assume it's from citizen (backward compatibility)
            citizen_files.append(archivo)
    
    if not citizen_files:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No hay archivos del ciudadano para descargar")
    
    import zipfile
    
    # Create ZIP file
    zip_filename = f"{petition['radicado']}_archivos_ciudadano.zip"
    zip_path = UPLOAD_DIR / zip_filename
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for archivo in citizen_files:
            file_path = Path(archivo['path'])
            if file_path.exists():
                # Add file to ZIP with original name
                zipf.write(file_path, archivo['original_name'])
    
    return FileResponse(
        path=zip_path,
        filename=zip_filename,
        media_type='application/zip'
    )

@api_router.post("/petitions/{petition_id}/assign-gestor")
async def assign_gestor(
    petition_id: str,
    assignment: GestorAssignment,
    current_user: dict = Depends(get_current_user)
):
    # Only atencion usuario, gestor, coordinador, and admin can assign
    if current_user['role'] not in [UserRole.ATENCION_USUARIO, UserRole.GESTOR, UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    gestor = await db.users.find_one({"id": assignment.gestor_id}, {"_id": 0})
    if not gestor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Gestor no encontrado")
    
    # Add gestor to assigned list
    gestores_asignados = petition.get('gestores_asignados', [])
    if assignment.gestor_id not in gestores_asignados:
        gestores_asignados.append(assignment.gestor_id)
    
    update_data = {
        "gestores_asignados": gestores_asignados,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # If first assignment, change status to ASIGNADO
    estado_cambio = False
    if petition['estado'] == PetitionStatus.RADICADO:
        update_data['estado'] = PetitionStatus.ASIGNADO
        estado_cambio = True
    
    # Add to historial
    gestor_rol = "Gestor" if gestor['role'] == UserRole.GESTOR else "Gestor Auxiliar"
    historial_entry = {
        "accion": f"{gestor_rol} asignado: {gestor['full_name']}",
        "usuario": current_user['full_name'],
        "usuario_rol": current_user['role'],
        "estado_anterior": petition['estado'],
        "estado_nuevo": update_data.get('estado', petition['estado']),
        "notas": f"Asignado a {gestor['full_name']} ({gestor_rol})",
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    current_historial = petition.get('historial', [])
    current_historial.append(historial_entry)
    update_data['historial'] = current_historial
    
    await db.petitions.update_one({"id": petition_id}, {"$set": update_data})
    
    # Send email to assigned gestor
    await send_email(
        gestor['email'],
        f"Trámite Asignado - {petition['radicado']}",
        f"<h3>Se te ha asignado un trámite</h3><p>Radicado: {petition['radicado']}</p><p>Tipo: {petition['tipo_tramite']}</p>"
    )
    
    return {"message": "Gestor asignado exitosamente"}

@api_router.patch("/petitions/{petition_id}")
async def update_petition(petition_id: str, update_data: PetitionUpdate, current_user: dict = Depends(get_current_user)):
    # Citizens cannot update petitions
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso para actualizar peticiones")
    
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Determine what fields can be updated based on role
    update_dict = {}
    historial_entry = None
    
    if current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        # Coordinador and Admin can update all fields
        update_dict = update_data.model_dump(exclude_none=True)
    elif current_user['role'] == UserRole.ATENCION_USUARIO:
        # Atención al usuario can update status, notes, and can finalize/reject
        if update_data.estado:
            update_dict['estado'] = update_data.estado
        if update_data.notas:
            update_dict['notas'] = update_data.notas
    elif current_user['role'] in [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]:
        # Gestores can only update notes and send to revision
        if update_data.notas:
            update_dict['notas'] = update_data.notas
        if update_data.estado in [PetitionStatus.REVISION, PetitionStatus.RECHAZADO]:
            update_dict['estado'] = update_data.estado
    
    if update_dict:
        update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        # Create historial entry if status changed
        if 'estado' in update_dict:
            estado_anterior = petition.get('estado')
            estado_nuevo = update_dict['estado']
            
            status_names = {
                PetitionStatus.RADICADO: "Radicado",
                PetitionStatus.ASIGNADO: "Asignado",
                PetitionStatus.RECHAZADO: "Rechazado",
                PetitionStatus.REVISION: "En Revisión",
                PetitionStatus.DEVUELTO: "Devuelto",
                PetitionStatus.FINALIZADO: "Finalizado"
            }
            
            historial_entry = {
                "accion": f"Estado cambiado de {status_names.get(estado_anterior, estado_anterior)} a {status_names.get(estado_nuevo, estado_nuevo)}",
                "usuario": current_user['full_name'],
                "usuario_rol": current_user['role'],
                "estado_anterior": estado_anterior,
                "estado_nuevo": estado_nuevo,
                "notas": update_dict.get('notas', ''),
                "fecha": datetime.now(timezone.utc).isoformat()
            }
            
            # Add to historial
            current_historial = petition.get('historial', [])
            current_historial.append(historial_entry)
            update_dict['historial'] = current_historial
        
        await db.petitions.update_one({"id": petition_id}, {"$set": update_dict})
        
        # Send email notification to citizen if status changed
        if 'estado' in update_dict:
            citizen = await db.users.find_one({"id": petition['user_id']}, {"_id": 0})
            if citizen:
                status_names = {
                    PetitionStatus.RADICADO: "Radicado",
                    PetitionStatus.ASIGNADO: "Asignado",
                    PetitionStatus.RECHAZADO: "Rechazado",
                    PetitionStatus.REVISION: "En Revisión",
                    PetitionStatus.DEVUELTO: "Devuelto",
                    PetitionStatus.FINALIZADO: "Finalizado"
                }
                await send_email(
                    citizen['email'],
                    f"Actualización de Trámite - {petition['radicado']}",
                    f"<h3>Su trámite ha sido actualizado</h3><p>Radicado: {petition['radicado']}</p><p>Nuevo estado: {status_names.get(update_dict['estado'])}</p>"
                )
    
    updated_petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if isinstance(updated_petition['created_at'], str):
        updated_petition['created_at'] = datetime.fromisoformat(updated_petition['created_at'])
    if isinstance(updated_petition['updated_at'], str):
        updated_petition['updated_at'] = datetime.fromisoformat(updated_petition['updated_at'])
    
    return updated_petition

@api_router.get("/petitions/stats/dashboard")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    if current_user['role'] == UserRole.CIUDADANO:
        query = {"user_id": current_user['id']}
    elif current_user['role'] in [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]:
        query = {"gestores_asignados": current_user['id']}
    else:
        query = {}
    
    total = await db.petitions.count_documents(query)
    radicado = await db.petitions.count_documents({**query, "estado": PetitionStatus.RADICADO})
    asignado = await db.petitions.count_documents({**query, "estado": PetitionStatus.ASIGNADO})
    rechazado = await db.petitions.count_documents({**query, "estado": PetitionStatus.RECHAZADO})
    revision = await db.petitions.count_documents({**query, "estado": PetitionStatus.REVISION})
    devuelto = await db.petitions.count_documents({**query, "estado": PetitionStatus.DEVUELTO})
    finalizado = await db.petitions.count_documents({**query, "estado": PetitionStatus.FINALIZADO})
    
    return {
        "total": total,
        "radicado": radicado,
        "asignado": asignado,
        "rechazado": rechazado,
        "revision": revision,
        "devuelto": devuelto,
        "finalizado": finalizado
    }

@api_router.get("/gestores")
async def get_gestores(current_user: dict = Depends(get_current_user)):
    # Get all gestores and auxiliares
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]}},
        {"_id": 0, "password": 0}
    ).to_list(1000)
    
    return gestores


# ===== PDF EXPORT AND DIGITAL SIGNATURE =====

def generate_petition_pdf(petition_data: dict, user_data: dict, signed_by: str = None) -> bytes:
    """Generate PDF report for a petition with optional digital signature"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#047857'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#064E3B'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12
    )
    
    # Title
    story.append(Paragraph("ASOCIACIÓN DE MUNICIPIOS DEL CATATUMBO", title_style))
    story.append(Paragraph("Provincia de Ocaña y Sur del Cesar - ASOMUNICIPIOS", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Radicado
    story.append(Paragraph(f"<b>Radicado:</b> {petition_data.get('radicado', 'N/A')}", heading_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Status
    status_names = {
        'radicado': 'Radicado',
        'asignado': 'Asignado',
        'rechazado': 'Rechazado',
        'revision': 'En Revisión',
        'devuelto': 'Devuelto',
        'finalizado': 'Finalizado'
    }
    status_label = status_names.get(petition_data.get('estado', ''), petition_data.get('estado', 'N/A'))
    story.append(Paragraph(f"<b>Estado:</b> {status_label}", normal_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Petition details table
    story.append(Paragraph("DATOS DEL SOLICITANTE", heading_style))
    
    data = [
        ['Campo', 'Información'],
        ['Nombre Completo', petition_data.get('nombre_completo', 'N/A')],
        ['Correo Electrónico', petition_data.get('correo', 'N/A')],
        ['Teléfono', petition_data.get('telefono', 'N/A')],
        ['Municipio', petition_data.get('municipio', 'N/A')],
        ['Tipo de Trámite', petition_data.get('tipo_tramite', 'N/A')],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#047857')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Notes if any
    if petition_data.get('notas'):
        story.append(Paragraph("NOTAS", heading_style))
        story.append(Paragraph(petition_data.get('notas', ''), normal_style))
        story.append(Spacer(1, 0.3*inch))
    
    # Dates
    story.append(Paragraph("INFORMACIÓN DE FECHAS", heading_style))
    created_date = petition_data.get('created_at', '')
    updated_date = petition_data.get('updated_at', '')
    
    if isinstance(created_date, str):
        created_date = datetime.fromisoformat(created_date).strftime('%d/%m/%Y %H:%M')
    else:
        created_date = created_date.strftime('%d/%m/%Y %H:%M') if created_date else 'N/A'
    
    if isinstance(updated_date, str):
        updated_date = datetime.fromisoformat(updated_date).strftime('%d/%m/%Y %H:%M')
    else:
        updated_date = updated_date.strftime('%d/%m/%Y %H:%M') if updated_date else 'N/A'
    
    story.append(Paragraph(f"<b>Fecha de Radicación:</b> {created_date}", normal_style))
    story.append(Paragraph(f"<b>Última Actualización:</b> {updated_date}", normal_style))
    story.append(Spacer(1, 0.5*inch))
    
    # Digital signature section
    if signed_by:
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph("___________________________", normal_style))
        story.append(Paragraph("<b>Firmado digitalmente por:</b>", normal_style))
        story.append(Paragraph(f"{signed_by}", normal_style))
        story.append(Paragraph(f"Fecha: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", normal_style))
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        "Este documento ha sido generado por el Sistema de Gestión Catastral de ASOMUNICIPIOS",
        footer_style
    ))
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


@api_router.get("/petitions/{petition_id}/export-pdf")
async def export_petition_pdf(petition_id: str, current_user: dict = Depends(get_current_user)):
    """Export single petition as PDF"""
    petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
    
    if not petition:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petición no encontrada")
    
    # Check permissions
    if current_user['role'] == UserRole.CIUDADANO and petition['user_id'] != current_user['id']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get user data
    user = await db.users.find_one({"id": petition['user_id']}, {"_id": 0, "password": 0})
    
    # Generate PDF with digital signature if coordinator or admin
    signed_by = None
    if current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        signed_by = f"{current_user['full_name']} - {current_user['role'].replace('_', ' ').title()}"
    
    pdf_bytes = generate_petition_pdf(petition, user, signed_by)
    
    # Save to temp file
    temp_pdf_path = UPLOAD_DIR / f"petition_{petition_id}_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"{petition['radicado']}.pdf",
        media_type='application/pdf'
    )


@api_router.post("/petitions/export-multiple")
async def export_multiple_petitions(
    petition_ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    """Export multiple petitions as PDF"""
    # Only staff can export multiple
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    from reportlab.platypus import PageBreak
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    
    for idx, petition_id in enumerate(petition_ids):
        petition = await db.petitions.find_one({"id": petition_id}, {"_id": 0})
        if petition:
            # Generate petition content
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#047857'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            
            story.append(Paragraph(f"Petición {idx + 1} de {len(petition_ids)}", title_style))
            story.append(Paragraph(f"Radicado: {petition.get('radicado', 'N/A')}", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            # Add basic info
            info = [
                ['Solicitante', petition.get('nombre_completo', 'N/A')],
                ['Tipo de Trámite', petition.get('tipo_tramite', 'N/A')],
                ['Estado', petition.get('estado', 'N/A')],
                ['Municipio', petition.get('municipio', 'N/A')],
            ]
            
            table = Table(info, colWidths=[2*inch, 4*inch])
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            story.append(table)
            
            # Add page break between petitions except for the last one
            if idx < len(petition_ids) - 1:
                story.append(PageBreak())
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    # Save to temp file
    temp_pdf_path = UPLOAD_DIR / f"petitions_report_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"reporte_peticiones_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


# ===== PRODUCTIVITY REPORTS =====

@api_router.get("/reports/gestor-productivity")
async def get_gestor_productivity(current_user: dict = Depends(get_current_user)):
    """Get productivity report for all gestores"""
    # Only admin, coordinador, and atencion_usuario can view reports
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get all gestores and auxiliares
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]}},
        {"_id": 0, "password": 0}
    ).to_list(1000)
    
    productivity_data = []
    
    for gestor in gestores:
        gestor_id = gestor['id']
        
        # Total petitions assigned
        total_assigned = await db.petitions.count_documents({
            "gestores_asignados": gestor_id
        })
        
        # Completed petitions
        completed = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.FINALIZADO
        })
        
        # In process
        in_process = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": {"$in": [PetitionStatus.ASIGNADO, PetitionStatus.REVISION, PetitionStatus.DEVUELTO]}
        })
        
        # Rejected
        rejected = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.RECHAZADO
        })
        
        # Calculate average time to complete (for completed petitions)
        completed_petitions = await db.petitions.find({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.FINALIZADO
        }, {"_id": 0, "created_at": 1, "updated_at": 1}).to_list(1000)
        
        avg_days = 0
        if completed_petitions:
            total_days = 0
            for petition in completed_petitions:
                created = petition['created_at']
                updated = petition['updated_at']
                
                if isinstance(created, str):
                    created = datetime.fromisoformat(created)
                if isinstance(updated, str):
                    updated = datetime.fromisoformat(updated)
                
                delta = (updated - created).days
                total_days += delta
            
            avg_days = round(total_days / len(completed_petitions), 1) if completed_petitions else 0
        
        # Calculate completion rate
        completion_rate = round((completed / total_assigned * 100), 1) if total_assigned > 0 else 0
        
        productivity_data.append({
            "gestor_id": gestor_id,
            "gestor_name": gestor['full_name'],
            "gestor_email": gestor['email'],
            "gestor_role": gestor['role'],
            "total_assigned": total_assigned,
            "completed": completed,
            "in_process": in_process,
            "rejected": rejected,
            "avg_completion_days": avg_days,
            "completion_rate": completion_rate
        })
    
    # Sort by completion rate descending
    productivity_data.sort(key=lambda x: x['completion_rate'], reverse=True)
    
    return productivity_data


@api_router.get("/reports/gestor-productivity/export-pdf")
async def export_gestor_productivity_pdf(current_user: dict = Depends(get_current_user)):
    """Export gestor productivity report as PDF"""
    # Only admin, coordinador, and atencion_usuario can export reports
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get productivity data
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]}},
        {"_id": 0, "password": 0}
    ).to_list(1000)
    
    productivity_data = []
    
    for gestor in gestores:
        gestor_id = gestor['id']
        
        total_assigned = await db.petitions.count_documents({"gestores_asignados": gestor_id})
        completed = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.FINALIZADO
        })
        in_process = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": {"$in": [PetitionStatus.ASIGNADO, PetitionStatus.REVISION, PetitionStatus.DEVUELTO]}
        })
        rejected = await db.petitions.count_documents({
            "gestores_asignados": gestor_id,
            "estado": PetitionStatus.RECHAZADO
        })
        
        completion_rate = round((completed / total_assigned * 100), 1) if total_assigned > 0 else 0
        
        productivity_data.append({
            "name": gestor['full_name'],
            "role": "Gestor" if gestor['role'] == UserRole.GESTOR else "Gestor Auxiliar",
            "total": total_assigned,
            "completed": completed,
            "in_process": in_process,
            "rejected": rejected,
            "rate": completion_rate
        })
    
    productivity_data.sort(key=lambda x: x['rate'], reverse=True)
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#047857'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph("ASOCIACIÓN DE MUNICIPIOS DEL CATATUMBO", title_style))
    story.append(Paragraph("Reporte de Productividad de Gestores", styles['Heading2']))
    story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 0.5*inch))
    
    # Table data
    table_data = [
        ['Gestor', 'Rol', 'Total', 'Finalizados', 'En Proceso', 'Rechazados', 'Tasa (%)']
    ]
    
    for data in productivity_data:
        table_data.append([
            data['name'],
            data['role'],
            str(data['total']),
            str(data['completed']),
            str(data['in_process']),
            str(data['rejected']),
            f"{data['rate']}%"
        ])
    
    # Create table
    col_widths = [1.8*inch, 1*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#047857')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.5*inch))
    
    # Summary
    story.append(Paragraph("Resumen General", styles['Heading3']))
    total_gestores = len(productivity_data)
    total_tramites = sum(d['total'] for d in productivity_data)
    total_finalizados = sum(d['completed'] for d in productivity_data)
    avg_rate = round(sum(d['rate'] for d in productivity_data) / total_gestores, 1) if total_gestores > 0 else 0
    
    summary_text = f"""
    <b>Total de Gestores:</b> {total_gestores}<br/>
    <b>Total de Trámites Asignados:</b> {total_tramites}<br/>
    <b>Total Finalizados:</b> {total_finalizados}<br/>
    <b>Tasa Promedio de Finalización:</b> {avg_rate}%
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        "Reporte generado por el Sistema de Gestión Catastral de ASOMUNICIPIOS",
        footer_style
    ))
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    # Save to temp file
    temp_pdf_path = UPLOAD_DIR / f"productivity_report_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"reporte_productividad_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


@api_router.get("/reports/listado-tramites/export-pdf")
async def export_listado_tramites_pdf(
    municipio: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export petition list as PDF (Listado de Trámites)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO, UserRole.GESTOR]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Build query
    query = {}
    if municipio:
        query["municipio"] = municipio
    if estado:
        query["estado"] = estado
    if fecha_inicio:
        query["created_at"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "created_at" in query:
            query["created_at"]["$lte"] = fecha_fin
        else:
            query["created_at"] = {"$lte": fecha_fin}
    
    # Get petitions
    petitions = await db.petitions.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Get user names for gestores
    users = await db.users.find({}, {"_id": 0, "id": 1, "full_name": 1}).to_list(1000)
    user_map = {u['id']: u['full_name'] for u in users}
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=0.5*inch, rightMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Header with logo
    logo_path = Path("/app/backend/logo_asomunicipios.png")
    if logo_path.exists():
        img = Image(str(logo_path), width=6*inch, height=1*inch)
        story.append(img)
        story.append(Spacer(1, 0.2*inch))
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#047857'),
        spaceAfter=10
    )
    story.append(Paragraph("LISTADO DE TRÁMITES CATASTRALES", title_style))
    
    # Filter info
    filter_text = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if municipio:
        filter_text += f" | Municipio: {municipio}"
    if estado:
        filter_text += f" | Estado: {estado}"
    story.append(Paragraph(filter_text, styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [
        ['No.', 'RADICADO', 'FECHA', 'SOLICITANTE', 'TIPO TRÁMITE', 'MUNICIPIO', 'ESTADO', 'GESTOR']
    ]
    
    estado_labels = {
        'radicado': 'Radicado',
        'asignado': 'Asignado',
        'revision': 'En Revisión',
        'devuelto': 'Devuelto',
        'rechazado': 'Rechazado',
        'finalizado': 'Finalizado'
    }
    
    for idx, pet in enumerate(petitions, 1):
        created_at = pet.get('created_at', '')
        if isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                fecha_str = created_at.strftime('%d/%m/%Y')
            except:
                fecha_str = str(created_at)[:10]
        else:
            fecha_str = created_at.strftime('%d/%m/%Y') if created_at else ''
        
        gestor_names = []
        for g_id in pet.get('gestores_asignados', []):
            if g_id in user_map:
                gestor_names.append(user_map[g_id])
        gestor_str = ', '.join(gestor_names) if gestor_names else 'Sin asignar'
        
        table_data.append([
            str(idx),
            pet.get('radicado_id', ''),
            fecha_str,
            pet.get('creator_name', ''),
            pet.get('tipo_tramite', ''),
            pet.get('municipio', ''),
            estado_labels.get(pet.get('estado', ''), pet.get('estado', '')),
            gestor_str[:30]  # Truncate long names
        ])
    
    # Create table
    col_widths = [0.4*inch, 1.3*inch, 0.8*inch, 1.8*inch, 1.5*inch, 1.2*inch, 0.9*inch, 1.5*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#047857')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9f4')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Summary
    total = len(petitions)
    finalizados = sum(1 for p in petitions if p.get('estado') == 'finalizado')
    en_proceso = sum(1 for p in petitions if p.get('estado') in ['asignado', 'revision', 'devuelto'])
    radicados = sum(1 for p in petitions if p.get('estado') == 'radicado')
    
    summary_text = f"<b>Resumen:</b> Total: {total} | Finalizados: {finalizados} | En Proceso: {en_proceso} | Radicados: {radicados}"
    story.append(Paragraph(summary_text, styles['Normal']))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(
        "ASOMUNICIPIOS - Asociación de Municipios del Catatumbo, Provincia de Ocaña y Sur del Cesar",
        footer_style
    ))
    
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    temp_pdf_path = UPLOAD_DIR / f"listado_tramites_{uuid.uuid4()}.pdf"
    with open(temp_pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=temp_pdf_path,
        filename=f"listado_tramites_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type='application/pdf'
    )


# ===== ADVANCED STATISTICS =====

@api_router.get("/stats/by-municipality")
async def get_stats_by_municipality(current_user: dict = Depends(get_current_user)):
    """Get petition statistics grouped by municipality"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {
            "_id": "$municipio",
            "total": {"$sum": 1},
            "radicado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RADICADO]}, 1, 0]}},
            "asignado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.ASIGNADO]}, 1, 0]}},
            "revision": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.REVISION]}, 1, 0]}},
            "finalizado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.FINALIZADO]}, 1, 0]}},
            "rechazado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RECHAZADO]}, 1, 0]}},
            "devuelto": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.DEVUELTO]}, 1, 0]}}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await db.petitions.aggregate(pipeline).to_list(100)
    
    return [
        {
            "municipio": r["_id"] or "Sin especificar",
            "total": r["total"],
            "radicado": r["radicado"],
            "asignado": r["asignado"],
            "revision": r["revision"],
            "finalizado": r["finalizado"],
            "rechazado": r["rechazado"],
            "devuelto": r["devuelto"]
        }
        for r in results
    ]

@api_router.get("/stats/by-tramite")
async def get_stats_by_tramite(current_user: dict = Depends(get_current_user)):
    """Get petition statistics grouped by tramite type"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {
            "_id": "$tipo_tramite",
            "total": {"$sum": 1},
            "radicado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RADICADO]}, 1, 0]}},
            "asignado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.ASIGNADO]}, 1, 0]}},
            "revision": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.REVISION]}, 1, 0]}},
            "finalizado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.FINALIZADO]}, 1, 0]}},
            "rechazado": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.RECHAZADO]}, 1, 0]}},
            "devuelto": {"$sum": {"$cond": [{"$eq": ["$estado", PetitionStatus.DEVUELTO]}, 1, 0]}}
        }},
        {"$sort": {"total": -1}}
    ]
    
    results = await db.petitions.aggregate(pipeline).to_list(100)
    
    return [
        {
            "tipo_tramite": r["_id"] or "Sin especificar",
            "total": r["total"],
            "radicado": r["radicado"],
            "asignado": r["asignado"],
            "revision": r["revision"],
            "finalizado": r["finalizado"],
            "rechazado": r["rechazado"],
            "devuelto": r["devuelto"]
        }
        for r in results
    ]

@api_router.get("/stats/by-gestor")
async def get_stats_by_gestor(current_user: dict = Depends(get_current_user)):
    """Get petition statistics grouped by assigned gestor"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Get all gestores
    gestores = await db.users.find(
        {"role": {"$in": [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]}},
        {"_id": 0, "id": 1, "full_name": 1, "role": 1}
    ).to_list(100)
    
    gestor_stats = []
    
    for gestor in gestores:
        gestor_id = gestor['id']
        
        # Count by status for this gestor
        total = await db.petitions.count_documents({"gestores_asignados": gestor_id})
        radicado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.RADICADO})
        asignado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.ASIGNADO})
        revision = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.REVISION})
        finalizado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.FINALIZADO})
        rechazado = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.RECHAZADO})
        devuelto = await db.petitions.count_documents({"gestores_asignados": gestor_id, "estado": PetitionStatus.DEVUELTO})
        
        completion_rate = round((finalizado / total * 100), 1) if total > 0 else 0
        
        gestor_stats.append({
            "gestor_id": gestor_id,
            "gestor_name": gestor['full_name'],
            "gestor_role": "Gestor" if gestor['role'] == UserRole.GESTOR else "Gestor Auxiliar",
            "total": total,
            "radicado": radicado,
            "asignado": asignado,
            "revision": revision,
            "finalizado": finalizado,
            "rechazado": rechazado,
            "devuelto": devuelto,
            "completion_rate": completion_rate
        })
    
    # Sort by total descending
    gestor_stats.sort(key=lambda x: x['total'], reverse=True)
    
    return gestor_stats

@api_router.get("/stats/summary")
async def get_stats_summary(current_user: dict = Depends(get_current_user)):
    """Get overall statistics summary"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tiene permiso")
    
    # Total counts
    total_petitions = await db.petitions.count_documents({})
    total_users = await db.users.count_documents({})
    total_gestores = await db.users.count_documents({"role": {"$in": [UserRole.GESTOR, UserRole.GESTOR_AUXILIAR]}})
    
    # Staff counts by role
    staff_counts = {
        "coordinadores": await db.users.count_documents({"role": UserRole.COORDINADOR}),
        "gestores": await db.users.count_documents({"role": UserRole.GESTOR}),
        "gestores_auxiliares": await db.users.count_documents({"role": UserRole.GESTOR_AUXILIAR}),
        "atencion_usuario": await db.users.count_documents({"role": UserRole.ATENCION_USUARIO}),
        "administradores": await db.users.count_documents({"role": UserRole.ADMINISTRADOR}),
        "ciudadanos": await db.users.count_documents({"role": UserRole.CIUDADANO})
    }
    
    # Status counts
    status_counts = {
        "radicado": await db.petitions.count_documents({"estado": PetitionStatus.RADICADO}),
        "asignado": await db.petitions.count_documents({"estado": PetitionStatus.ASIGNADO}),
        "revision": await db.petitions.count_documents({"estado": PetitionStatus.REVISION}),
        "finalizado": await db.petitions.count_documents({"estado": PetitionStatus.FINALIZADO}),
        "rechazado": await db.petitions.count_documents({"estado": PetitionStatus.RECHAZADO}),
        "devuelto": await db.petitions.count_documents({"estado": PetitionStatus.DEVUELTO})
    }
    
    # Completion rate
    completion_rate = round((status_counts["finalizado"] / total_petitions * 100), 1) if total_petitions > 0 else 0
    
    # Recent petitions (last 30 days)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    recent_petitions = await db.petitions.count_documents({
        "created_at": {"$gte": thirty_days_ago}
    })
    
    return {
        "total_petitions": total_petitions,
        "total_users": total_users,
        "total_gestores": total_gestores,
        "staff_counts": staff_counts,
        "status_counts": status_counts,
        "completion_rate": completion_rate,
        "recent_petitions_30_days": recent_petitions
    }


# ===== PREDIOS ROUTES (Código Nacional Catastral) =====

async def generate_codigo_predial(municipio: str, zona: str, sector: str, manzana_vereda: str, 
                                   terreno: str, condicion: str, ph: str) -> str:
    """Genera el código predial nacional de 30 dígitos"""
    if municipio not in MUNICIPIOS_DIVIPOLA:
        raise HTTPException(status_code=400, detail=f"Municipio '{municipio}' no válido")
    
    divipola = MUNICIPIOS_DIVIPOLA[municipio]
    
    # Construir código de 30 dígitos
    codigo = (
        divipola["departamento"].zfill(2) +  # 2 dígitos
        divipola["municipio"].zfill(3) +     # 3 dígitos
        zona.zfill(2) +                       # 2 dígitos
        sector.zfill(2) +                     # 2 dígitos
        manzana_vereda.zfill(4) +            # 4 dígitos
        terreno.zfill(4) +                    # 4 dígitos
        condicion.zfill(4) +                  # 4 dígitos
        ph.zfill(4) +                         # 4 dígitos
        "00000"                               # 5 dígitos (unidad predial)
    )
    
    return codigo

async def generate_codigo_homologado(municipio: str) -> str:
    """Genera un código homologado único de 11 caracteres"""
    import string
    import random
    
    # Obtener último código para este municipio
    last_predio = await db.predios.find_one(
        {"municipio": municipio, "deleted": {"$ne": True}},
        sort=[("numero_predio", -1)]
    )
    
    if last_predio:
        next_num = last_predio.get("numero_predio", 0) + 1
    else:
        next_num = 1
    
    # Generar código: BPP + número + letras aleatorias
    letters = ''.join(random.choices(string.ascii_uppercase, k=4))
    codigo = f"BPP{str(next_num).zfill(4)}{letters}"
    
    return codigo, next_num

async def get_next_terreno_number(municipio: str, zona: str, sector: str, manzana_vereda: str) -> str:
    """Obtiene el siguiente número de terreno disponible (incluyendo eliminados)"""
    # Buscar el máximo terreno usado (incluyendo eliminados para no reutilizar)
    pipeline = [
        {"$match": {
            "municipio": municipio,
            "zona": zona,
            "sector": sector,
            "manzana_vereda": manzana_vereda
        }},
        {"$group": {
            "_id": None,
            "max_terreno": {"$max": "$terreno_num"}
        }}
    ]
    
    result = await db.predios.aggregate(pipeline).to_list(1)
    
    if result and result[0].get("max_terreno"):
        next_num = result[0]["max_terreno"] + 1
    else:
        next_num = 1
    
    return str(next_num).zfill(4), next_num

@api_router.get("/predios/catalogos")
async def get_predios_catalogos(current_user: dict = Depends(get_current_user)):
    """Obtiene los catálogos para el formulario de predios"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    return {
        "municipios": list(MUNICIPIOS_DIVIPOLA.keys()),
        "destino_economico": DESTINO_ECONOMICO,
        "tipo_documento": TIPO_DOCUMENTO_PREDIO,
        "estado_civil": ESTADO_CIVIL_PREDIO,
        "divipola": MUNICIPIOS_DIVIPOLA
    }

@api_router.get("/predios")
async def get_predios(
    municipio: Optional[str] = None,
    vigencia: Optional[int] = None,
    destino_economico: Optional[str] = None,
    zona: Optional[str] = None,  # '00' = urbano, otros = rural
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista todos los predios (solo staff)"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"deleted": {"$ne": True}}
    
    if municipio:
        query["municipio"] = municipio
    if vigencia:
        query["vigencia"] = vigencia
    if destino_economico:
        query["destino_economico"] = destino_economico
    if zona:
        if zona == 'urbano':
            query["zona"] = "00"
        elif zona == 'rural':
            query["zona"] = {"$ne": "00"}
    if search:
        query["$or"] = [
            {"codigo_predial_nacional": {"$regex": search, "$options": "i"}},
            {"codigo_homologado": {"$regex": search, "$options": "i"}},
            {"propietarios.nombre_propietario": {"$regex": search, "$options": "i"}},
            {"propietarios.numero_documento": {"$regex": search, "$options": "i"}},
            {"direccion": {"$regex": search, "$options": "i"}},
            {"matriculas": {"$regex": search, "$options": "i"}}  # Búsqueda por matrícula
        ]
    
    total = await db.predios.count_documents(query)
    predios = await db.predios.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "predios": predios
    }

@api_router.get("/predios/stats/summary")
async def get_predios_stats(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas de predios - SOLO la vigencia más alta GLOBAL del sistema"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Función para extraer el año de una vigencia
    def get_year(vig):
        vig_str = str(vig)
        if len(vig_str) >= 7:
            return int(vig_str[-4:])
        return int(vig_str)
    
    # Obtener TODAS las vigencias disponibles en el sistema
    all_vigencias = await db.predios.distinct("vigencia", {"deleted": {"$ne": True}})
    
    if not all_vigencias:
        return {
            "total_predios": 0,
            "total_avaluo": 0,
            "total_area_terreno": 0,
            "by_municipio": [],
            "by_destino": [],
            "vigencia_actual": None
        }
    
    # Encontrar la vigencia más alta (año más reciente) GLOBALMENTE
    vigencia_mas_alta = max(all_vigencias, key=lambda x: get_year(x))
    vigencia_year = get_year(vigencia_mas_alta)
    
    # Filtrar solo predios de la vigencia más alta
    pipeline_municipios = [
        {"$match": {"vigencia": vigencia_mas_alta, "deleted": {"$ne": True}}},
        {"$group": {"_id": "$municipio", "count": {"$sum": 1}, "avaluo": {"$sum": "$avaluo"}, "area": {"$sum": "$area_terreno"}}},
        {"$sort": {"count": -1}}
    ]
    
    municipios_result = await db.predios.aggregate(pipeline_municipios).to_list(100)
    
    by_municipio = []
    total_predios = 0
    total_avaluo = 0
    total_area = 0
    
    for r in municipios_result:
        by_municipio.append({
            "municipio": r['_id'], 
            "count": r['count'],
            "vigencia": vigencia_mas_alta,
            "vigencia_display": str(vigencia_year)
        })
        total_predios += r['count']
        total_avaluo += r.get('avaluo', 0) or 0
        total_area += r.get('area', 0) or 0
    
    # Por destino económico (solo vigencia más alta)
    pipeline_destino = [
        {"$match": {"vigencia": vigencia_mas_alta, "deleted": {"$ne": True}}},
        {"$group": {"_id": "$destino_economico", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    destinos_result = await db.predios.aggregate(pipeline_destino).to_list(20)
    
    by_destino = [
        {"destino": d["_id"], "nombre": DESTINO_ECONOMICO.get(d["_id"], "Desconocido"), "count": d["count"]}
        for d in destinos_result
    ]
    
    return {
        "total_predios": total_predios,
        "total_avaluo": total_avaluo,
        "total_area_terreno": total_area,
        "by_municipio": by_municipio,
        "by_destino": by_destino[:20],
        "vigencia_actual": vigencia_year
    }

@api_router.get("/predios/eliminados")
async def get_predios_eliminados(
    municipio: Optional[str] = None,
    vigencia: Optional[int] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista predios eliminados - filtrable por municipio y vigencia"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {}
    if municipio:
        query["municipio"] = municipio
    if vigencia:
        query["vigencia_eliminacion"] = vigencia
    
    total = await db.predios_eliminados.count_documents(query)
    predios = await db.predios_eliminados.find(query, {"_id": 0}).sort("eliminado_en", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "predios": predios
    }

@api_router.get("/predios/eliminados/stats")
async def get_predios_eliminados_stats(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas de predios eliminados por municipio"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    pipeline = [
        {"$group": {"_id": {"municipio": "$municipio", "vigencia": "$vigencia_eliminacion"}, "count": {"$sum": 1}}},
        {"$sort": {"_id.municipio": 1}}
    ]
    
    result = await db.predios_eliminados.aggregate(pipeline).to_list(100)
    
    return {
        "by_municipio": [
            {"municipio": r["_id"]["municipio"], "vigencia": r["_id"]["vigencia"], "count": r["count"]}
            for r in result
        ],
        "total": sum(r["count"] for r in result)
    }


@api_router.post("/predios/import-excel")
async def import_predios_excel(
    file: UploadFile = File(...),
    vigencia: int = 2025,
    current_user: dict = Depends(get_current_user)
):
    """Importa predios desde archivo Excel R1-R2 con soporte de vigencia"""
    import openpyxl
    
    # Helper para convertir números con formato de coma decimal
    def parse_number(value, default=0):
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        try:
            # Convertir string, reemplazando coma por punto
            s = str(value).strip().replace('$', '').replace(' ', '')
            # Si tiene punto y coma, el punto es separador de miles
            if '.' in s and ',' in s:
                s = s.replace('.', '').replace(',', '.')
            elif ',' in s:
                s = s.replace(',', '.')
            return float(s) if s else default
        except:
            return default
    
    # Solo coordinador o admin pueden importar
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden importar datos")
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")
    
    try:
        # Guardar archivo temporalmente
        temp_path = UPLOAD_DIR / f"temp_import_{uuid.uuid4()}.xlsx"
        content = await file.read()
        with open(temp_path, 'wb') as f:
            f.write(content)
        
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
        
        # Buscar hoja R1 con nombres alternativos (normalizando espacios)
        r1_sheet_names = ['REGISTRO_R1', 'R1', 'Registro_R1', 'registro_r1', 'Hoja1', 'Sheet1']
        ws_r1 = None
        # Primero intentar coincidencia exacta
        for name in r1_sheet_names:
            if name in wb.sheetnames:
                ws_r1 = wb[name]
                break
        # Si no encontró, intentar con nombres que contengan espacios
        if ws_r1 is None:
            for sheet_name in wb.sheetnames:
                sheet_name_clean = sheet_name.strip().upper()
                for name in r1_sheet_names:
                    if sheet_name_clean == name.upper():
                        ws_r1 = wb[sheet_name]
                        break
                if ws_r1:
                    break
        
        if ws_r1 is None:
            wb.close()
            temp_path.unlink()
            raise HTTPException(
                status_code=400, 
                detail=f"No se encontró hoja R1. Hojas disponibles: {', '.join(wb.sheetnames)}. Se esperaba: REGISTRO_R1, R1, o similar."
            )
        
        # Leer R1 (propietarios)
        r1_data = {}
        rows_read = 0
        
        for row in ws_r1.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Sin departamento = fila vacía
                continue
            
            rows_read += 1
            codigo_predial = str(row[3] or '').strip()
            if not codigo_predial:
                continue
            
            if codigo_predial not in r1_data:
                r1_data[codigo_predial] = {
                    'departamento': str(row[0] or '').strip(),
                    'municipio': str(row[1] or '').strip(),
                    'numero_predio': str(row[2] or '').strip(),
                    'codigo_predial_nacional': codigo_predial,
                    'codigo_homologado': str(row[4] or '').strip(),
                    'direccion': str(row[12] or '').strip(),
                    'comuna': str(row[13] or '').strip(),
                    'destino_economico': str(row[14] or '').strip(),
                    'area_terreno': parse_number(row[15]),
                    'area_construida': parse_number(row[16]),
                    'avaluo': parse_number(row[17]),
                    'vigencia': vigencia,
                    'tipo_mutacion': str(row[19] or '').strip() if len(row) > 19 else '',
                    'numero_resolucion': str(row[20] or '').strip() if len(row) > 20 else '',
                    'fecha_resolucion': str(row[21] or '').strip() if len(row) > 21 else '',
                    'propietarios': [],
                    'r2_registros': [],
                    'id': str(uuid.uuid4()),
                    'created_at': datetime.now(timezone.utc).isoformat(),
                    'status': 'aprobado'
                }
            
            # Agregar propietario
            nombre = str(row[8] or '').strip()
            if nombre:
                r1_data[codigo_predial]['propietarios'].append({
                    'nombre_propietario': nombre,
                    'estado_civil': str(row[9] or '').strip(),
                    'tipo_documento': str(row[10] or '').strip(),
                    'numero_documento': str(row[11] or '').strip()
                })
        
        # Buscar hoja R2 con nombres alternativos (normalizando espacios)
        r2_sheet_names = ['REGISTRO_R2', 'R2', 'Registro_R2', 'registro_r2', 'Hoja2', 'Sheet2']
        ws_r2 = None
        # Primero intentar coincidencia exacta
        for name in r2_sheet_names:
            if name in wb.sheetnames:
                ws_r2 = wb[name]
                break
        # Si no encontró, intentar con nombres que contengan espacios
        if ws_r2 is None:
            for sheet_name in wb.sheetnames:
                sheet_name_clean = sheet_name.strip().upper()
                for name in r2_sheet_names:
                    if sheet_name_clean == name.upper():
                        ws_r2 = wb[sheet_name]
                        break
                if ws_r2:
                    break
        
        if ws_r2 is None:
            wb.close()
            temp_path.unlink()
            raise HTTPException(
                status_code=400, 
                detail=f"No se encontró hoja R2. Hojas disponibles: {', '.join(wb.sheetnames)}. Se esperaba: REGISTRO_R2, R2, o similar."
            )
        
        # Leer R2 (físico)
        for row in ws_r2.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            codigo_predial = str(row[3] or '').strip()
            if codigo_predial not in r1_data:
                continue
            
            matricula = str(row[7] or '').strip() if len(row) > 7 else ''
            
            # Buscar si ya existe este registro R2 (por matrícula o agregar nuevo)
            r2_exists = False
            for r2 in r1_data[codigo_predial]['r2_registros']:
                if r2.get('matricula_inmobiliaria') == matricula:
                    r2_exists = True
                    break
            
            if not r2_exists:
                zonas = []
                
                # Zona 1
                if len(row) > 10 and row[10]:
                    area_t = parse_number(row[10])
                    zonas.append({
                        'zona_fisica': str(row[8] or '').strip() if len(row) > 8 else '',
                        'zona_economica': str(row[9] or '').strip() if len(row) > 9 else '',
                        'area_terreno': area_t,
                        'habitaciones': int(parse_number(row[14])) if len(row) > 14 else 0,
                        'banos': int(parse_number(row[15])) if len(row) > 15 else 0,
                        'locales': int(parse_number(row[16])) if len(row) > 16 else 0,
                        'pisos': int(parse_number(row[17])) if len(row) > 17 else 0,
                        'tipificacion': str(row[18] or '').strip() if len(row) > 18 else '',
                        'uso': str(row[19] or '').strip() if len(row) > 19 else '',
                        'puntaje': int(parse_number(row[20])) if len(row) > 20 else 0,
                        'area_construida': parse_number(row[21]) if len(row) > 21 else 0
                    })
                
                # Zona 2
                if len(row) > 13 and row[13]:
                    area_t2 = parse_number(row[13])
                    if area_t2 > 0:
                        zonas.append({
                            'zona_fisica': str(row[11] or '').strip() if len(row) > 11 else '',
                            'zona_economica': str(row[12] or '').strip() if len(row) > 12 else '',
                            'area_terreno': area_t2,
                            'habitaciones': int(parse_number(row[22])) if len(row) > 22 else 0,
                            'banos': int(parse_number(row[23])) if len(row) > 23 else 0,
                            'locales': int(parse_number(row[24])) if len(row) > 24 else 0,
                            'pisos': int(parse_number(row[25])) if len(row) > 25 else 0,
                            'tipificacion': str(row[26] or '').strip() if len(row) > 26 else '',
                            'uso': str(row[27] or '').strip() if len(row) > 27 else '',
                            'puntaje': int(parse_number(row[28])) if len(row) > 28 else 0,
                            'area_construida': parse_number(row[29]) if len(row) > 29 else 0
                        })
                
                r1_data[codigo_predial]['r2_registros'].append({
                    'matricula_inmobiliaria': matricula,
                    'zonas': zonas
                })
        
        wb.close()
        temp_path.unlink()
        
        # Mapeo de códigos de municipio a nombres
        MUNICIPIO_CODIGOS = {
            '003': 'Ábrego', '54003': 'Ábrego', '3': 'Ábrego',
            '109': 'Bucarasica', '54109': 'Bucarasica',
            '128': 'Cáchira', '54128': 'Cáchira',
            '206': 'Convención', '54206': 'Convención',
            '245': 'El Carmen', '54245': 'El Carmen',
            '250': 'El Tarra', '54250': 'El Tarra',
            '344': 'Hacarí', '54344': 'Hacarí',
            '398': 'La Playa', '54398': 'La Playa',
            '670': 'San Calixto', '54670': 'San Calixto',
            '720': 'Sardinata', '54720': 'Sardinata',
            '800': 'Teorama', '54800': 'Teorama',
            '614': 'Río de Oro', '20614': 'Río de Oro',
        }
        
        # Función para extraer municipio del código predial nacional
        def get_municipio_from_codigo(codigo_predial):
            """Extrae el código de municipio del código predial nacional.
            Formato: DDMMMXXXX... donde DD=depto, MMM=municipio
            """
            if not codigo_predial or len(codigo_predial) < 5:
                return None
            # Los primeros 2 dígitos son departamento, los siguientes 3 son municipio
            codigo_mun = codigo_predial[2:5]
            # Intentar con el código completo depto+mun
            codigo_completo = codigo_predial[:5]
            if codigo_completo in MUNICIPIO_CODIGOS:
                return MUNICIPIO_CODIGOS[codigo_completo]
            if codigo_mun in MUNICIPIO_CODIGOS:
                return MUNICIPIO_CODIGOS[codigo_mun]
            # Intentar sin ceros a la izquierda
            codigo_mun_int = codigo_mun.lstrip('0')
            if codigo_mun_int in MUNICIPIO_CODIGOS:
                return MUNICIPIO_CODIGOS[codigo_mun_int]
            return None
        
        # Guardar en historial antes de actualizar
        municipio_raw = list(r1_data.values())[0]['municipio'] if r1_data else 'Desconocido'
        primer_codigo = list(r1_data.keys())[0] if r1_data else ''
        
        # Primero intentar extraer del código predial nacional (más confiable)
        municipio = get_municipio_from_codigo(primer_codigo)
        
        # Si no funciona, intentar con el valor raw
        if not municipio:
            municipio = MUNICIPIO_CODIGOS.get(municipio_raw, None)
        
        # Si el municipio_raw es un código predial largo, intentar extraerlo también
        if not municipio and len(str(municipio_raw)) >= 5:
            municipio = get_municipio_from_codigo(municipio_raw)
        
        if not municipio:
            # Usar el valor raw como último recurso
            municipio = municipio_raw
            logger.warning(f"No se pudo determinar el municipio. raw={municipio_raw}, codigo={primer_codigo[:30]}")
        
        # Actualizar el municipio en todos los predios
        for predio in r1_data.values():
            predio['municipio'] = municipio
        
        # Obtener los códigos prediales de los nuevos predios
        new_codigos = set(r1_data.keys())
        
        # Obtener los predios existentes del municipio PARA ESTA VIGENCIA ESPECÍFICA
        existing_predios = await db.predios.find(
            {"municipio": municipio, "vigencia": vigencia}, 
            {"_id": 0}
        ).to_list(50000)
        
        existing_codigos = {p.get('codigo_predial_nacional') for p in existing_predios}
        
        # Calcular predios eliminados (estaban antes en esta vigencia pero no vienen en la nueva importación)
        codigos_eliminados = existing_codigos - new_codigos
        predios_eliminados_count = len(codigos_eliminados)
        
        # Guardar predios eliminados en colección separada
        if codigos_eliminados:
            predios_a_eliminar = [p for p in existing_predios if p.get('codigo_predial_nacional') in codigos_eliminados]
            if predios_a_eliminar:
                await db.predios_eliminados.insert_many([
                    {
                        **p, 
                        "eliminado_en": datetime.now(timezone.utc).isoformat(),
                        "vigencia_eliminacion": vigencia,
                        "motivo": "No incluido en nueva importación R1-R2"
                    }
                    for p in predios_a_eliminar
                ])
        
        # Guardar historial de predios existentes de esta vigencia
        if existing_predios:
            await db.predios_historico.insert_many([
                {**p, "archivado_en": datetime.now(timezone.utc).isoformat(), "vigencia_archivo": vigencia}
                for p in existing_predios
            ])
        
        # Eliminar predios actuales de este municipio SOLO PARA ESTA VIGENCIA
        await db.predios.delete_many({"municipio": municipio, "vigencia": vigencia})
        
        # Insertar nuevos predios
        predios_list = list(r1_data.values())
        if predios_list:
            await db.predios.insert_many(predios_list)
        
        # Calcular predios nuevos (no estaban antes)
        predios_nuevos_count = len(new_codigos - existing_codigos)
        
        # Registrar importación
        logger.info(f"Import stats: rows_read={rows_read}, unique_predios={len(r1_data)}, municipio={municipio}")
        await db.importaciones.insert_one({
            "id": str(uuid.uuid4()),
            "municipio": municipio,
            "vigencia": vigencia,
            "total_predios": len(predios_list),
            "predios_anteriores": len(existing_predios),
            "predios_eliminados": predios_eliminados_count,
            "predios_nuevos": predios_nuevos_count,
            "archivo": file.filename,
            "importado_por": current_user['id'],
            "importado_por_nombre": current_user['full_name'],
            "fecha": datetime.now(timezone.utc).isoformat()
        })
        
        return {
            "message": f"Importación exitosa para {municipio}",
            "vigencia": vigencia,
            "predios_importados": len(predios_list),
            "predios_anteriores": len(existing_predios),
            "predios_eliminados": predios_eliminados_count,
            "predios_nuevos": predios_nuevos_count,
            "municipio": municipio
        }
        
    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error al importar: {str(e)}")


@api_router.get("/predios/vigencias")
async def get_vigencias_disponibles(current_user: dict = Depends(get_current_user)):
    """Obtiene las vigencias disponibles por municipio, ordenadas de más reciente a más antigua"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Vigencias actuales
    pipeline = [
        {"$group": {"_id": {"municipio": "$municipio", "vigencia": "$vigencia"}, "count": {"$sum": 1}}},
        {"$sort": {"_id.municipio": 1}}
    ]
    result = await db.predios.aggregate(pipeline).to_list(1000)
    
    vigencias = {}
    for r in result:
        mun = r['_id']['municipio']
        vig = r['_id']['vigencia']
        if mun not in vigencias:
            vigencias[mun] = []
        vigencias[mun].append({"vigencia": vig, "predios": r['count']})
    
    # Vigencias históricas
    historico_pipeline = [
        {"$group": {"_id": {"municipio": "$municipio", "vigencia": "$vigencia"}, "count": {"$sum": 1}}},
        {"$sort": {"_id.municipio": 1}}
    ]
    historico = await db.predios_historico.aggregate(historico_pipeline).to_list(1000)
    
    for h in historico:
        mun = h['_id']['municipio']
        vig = h['_id']['vigencia']
        if mun not in vigencias:
            vigencias[mun] = []
        # Agregar si no existe
        if not any(v['vigencia'] == vig for v in vigencias[mun]):
            vigencias[mun].append({"vigencia": vig, "predios": h['count'], "historico": True})
    
    # Función para extraer el año de una vigencia (puede ser 2025, 01012025, 1012025)
    def get_year(vig):
        vig_str = str(vig)
        if len(vig_str) >= 7:
            return int(vig_str[-4:])
        return int(vig_str)
    
    # Ordenar vigencias: primero datos actuales (no históricos) por año descendente, luego históricos
    for mun in vigencias:
        # Separar actuales de históricos
        actuales = [v for v in vigencias[mun] if not v.get('historico')]
        historicos = [v for v in vigencias[mun] if v.get('historico')]
        
        # Ordenar cada grupo por año descendente
        actuales.sort(key=lambda x: get_year(x['vigencia']), reverse=True)
        historicos.sort(key=lambda x: get_year(x['vigencia']), reverse=True)
        
        # Combinar: actuales primero, luego históricos
        vigencias[mun] = actuales + historicos
    
    return vigencias


@api_router.get("/predios/export-excel")
async def export_predios_excel(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Exporta predios a Excel en formato EXACTO al archivo original R1-R2"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Query
    query = {"deleted": {"$ne": True}}
    if municipio:
        query["municipio"] = municipio
    
    predios = await db.predios.find(query, {"_id": 0}).to_list(50000)
    
    # Crear workbook
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === HOJA REGISTRO_R1 (Propietarios) ===
    ws_r1 = wb.active
    ws_r1.title = "REGISTRO_R1"
    
    # Headers R1 - EXACTO al original
    headers_r1 = [
        "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL", 
        "CODIGO_HOMOLOGADO", "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS",
        "NOMBRE", "ESTADO_CIVIL", "TIPO_DOCUMENTO", "NUMERO_DOCUMENTO", "DIRECCION",
        "COMUNA", "DESTINO_ECONOMICO", "AREA_TERRENO", "AREA_CONSTRUIDA", "AVALUO",
        "VIGENCIA", "TIPO_MUTACIÓN", "NO. RESOLUCIÓN", "FECHA_RESOLUCIÓN"
    ]
    
    for col, header in enumerate(headers_r1, 1):
        cell = ws_r1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos R1
    row = 2
    for predio in predios:
        propietarios = predio.get('propietarios', [])
        if not propietarios:
            propietarios = [{'nombre_propietario': predio.get('nombre_propietario', ''),
                           'tipo_documento': predio.get('tipo_documento', ''),
                           'numero_documento': predio.get('numero_documento', ''),
                           'estado_civil': predio.get('estado_civil', '')}]
        
        total_props = len(propietarios)
        for idx, prop in enumerate(propietarios, 1):
            ws_r1.cell(row=row, column=1, value=predio.get('departamento', ''))
            ws_r1.cell(row=row, column=2, value=predio.get('municipio', ''))
            ws_r1.cell(row=row, column=3, value=predio.get('numero_predio', ''))
            ws_r1.cell(row=row, column=4, value=predio.get('codigo_predial_nacional', ''))
            ws_r1.cell(row=row, column=5, value=predio.get('codigo_homologado', ''))
            ws_r1.cell(row=row, column=6, value='1')
            ws_r1.cell(row=row, column=7, value=str(idx).zfill(2))
            ws_r1.cell(row=row, column=8, value=str(total_props).zfill(2))
            ws_r1.cell(row=row, column=9, value=prop.get('nombre_propietario', ''))
            ws_r1.cell(row=row, column=10, value=prop.get('estado_civil', ''))
            ws_r1.cell(row=row, column=11, value=prop.get('tipo_documento', ''))
            ws_r1.cell(row=row, column=12, value=prop.get('numero_documento', ''))
            ws_r1.cell(row=row, column=13, value=predio.get('direccion', ''))
            ws_r1.cell(row=row, column=14, value=predio.get('comuna', ''))
            ws_r1.cell(row=row, column=15, value=predio.get('destino_economico', ''))
            ws_r1.cell(row=row, column=16, value=predio.get('area_terreno', 0))
            ws_r1.cell(row=row, column=17, value=predio.get('area_construida', 0))
            ws_r1.cell(row=row, column=18, value=predio.get('avaluo', 0))
            ws_r1.cell(row=row, column=19, value=predio.get('vigencia', datetime.now().year))
            ws_r1.cell(row=row, column=20, value=predio.get('tipo_mutacion', ''))
            ws_r1.cell(row=row, column=21, value=predio.get('numero_resolucion', ''))
            ws_r1.cell(row=row, column=22, value=predio.get('fecha_resolucion', ''))
            row += 1
    
    # === HOJA REGISTRO_R2 (Físico - con zonas en columnas horizontales) ===
    ws_r2 = wb.create_sheet(title="REGISTRO_R2")
    
    # Headers R2 - EXACTO al original con zonas en columnas horizontales
    headers_r2 = [
        "DEPARTAMENTO", "MUNICIPIO", "NUMERO_DEL_PREDIO", "CODIGO_PREDIAL_NACIONAL",
        "TIPO_DE_REGISTRO", "NUMERO_DE_ORDEN", "TOTAL_REGISTROS", "MATRICULA_INMOBILIARIA",
        # Zona 1
        "ZONA_FISICA_1", "ZONA_ECONOMICA_1", "AREA_TERRENO_1",
        # Zona 2
        "ZONA_FISICA_2", "ZONA_ECONOMICA_2", "AREA_TERRENO_2",
        # Construcción 1
        "HABITACIONES_1", "BANOS_1", "LOCALES_1", "PISOS_1", "TIPIFICACION_1", "USO_1", "PUNTAJE_1", "AREA_CONSTRUIDA_1",
        # Construcción 2
        "HABITACIONES_2", "BANOS_2", "LOCALES_2", "PISOS_2", "TIPIFICACION_2", "USO_2", "PUNTAJE_2", "AREA_CONSTRUIDA_2",
        # Construcción 3
        "HABITACIONES_3", "BANOS_3", "LOCALES_3", "PISOS_3", "TIPIFICACION_3", "USO_3", "PUNTAJE_3", "AREA_CONSTRUIDA_3",
        "VIGENCIA"
    ]
    
    for col, header in enumerate(headers_r2, 1):
        cell = ws_r2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos R2 - Una fila por registro R2 con zonas en columnas
    row = 2
    for predio in predios:
        r2_registros = predio.get('r2_registros', [])
        total_r2 = len(r2_registros) if r2_registros else 0
        
        for r2_idx, r2 in enumerate(r2_registros, 1):
            ws_r2.cell(row=row, column=1, value=predio.get('departamento', ''))
            ws_r2.cell(row=row, column=2, value=predio.get('municipio', ''))
            ws_r2.cell(row=row, column=3, value=predio.get('numero_predio', ''))
            ws_r2.cell(row=row, column=4, value=predio.get('codigo_predial_nacional', ''))
            ws_r2.cell(row=row, column=5, value='2')
            ws_r2.cell(row=row, column=6, value=str(r2_idx).zfill(2))
            ws_r2.cell(row=row, column=7, value=str(total_r2).zfill(2))
            ws_r2.cell(row=row, column=8, value=r2.get('matricula_inmobiliaria', ''))
            
            zonas = r2.get('zonas', [])
            
            # Zona 1 (columnas 9-11)
            if len(zonas) >= 1:
                ws_r2.cell(row=row, column=9, value=zonas[0].get('zona_fisica', ''))
                ws_r2.cell(row=row, column=10, value=zonas[0].get('zona_economica', ''))
                ws_r2.cell(row=row, column=11, value=zonas[0].get('area_terreno', 0))
            
            # Zona 2 (columnas 12-14)
            if len(zonas) >= 2:
                ws_r2.cell(row=row, column=12, value=zonas[1].get('zona_fisica', ''))
                ws_r2.cell(row=row, column=13, value=zonas[1].get('zona_economica', ''))
                ws_r2.cell(row=row, column=14, value=zonas[1].get('area_terreno', 0))
            
            # Construcción 1 (columnas 15-22)
            if len(zonas) >= 1:
                ws_r2.cell(row=row, column=15, value=zonas[0].get('habitaciones', 0))
                ws_r2.cell(row=row, column=16, value=zonas[0].get('banos', 0))
                ws_r2.cell(row=row, column=17, value=zonas[0].get('locales', 0))
                ws_r2.cell(row=row, column=18, value=zonas[0].get('pisos', 0))
                ws_r2.cell(row=row, column=19, value=zonas[0].get('tipificacion', ''))
                ws_r2.cell(row=row, column=20, value=zonas[0].get('uso', ''))
                ws_r2.cell(row=row, column=21, value=zonas[0].get('puntaje', 0))
                ws_r2.cell(row=row, column=22, value=zonas[0].get('area_construida', 0))
            
            # Construcción 2 (columnas 23-30)
            if len(zonas) >= 2:
                ws_r2.cell(row=row, column=23, value=zonas[1].get('habitaciones', 0))
                ws_r2.cell(row=row, column=24, value=zonas[1].get('banos', 0))
                ws_r2.cell(row=row, column=25, value=zonas[1].get('locales', 0))
                ws_r2.cell(row=row, column=26, value=zonas[1].get('pisos', 0))
                ws_r2.cell(row=row, column=27, value=zonas[1].get('tipificacion', ''))
                ws_r2.cell(row=row, column=28, value=zonas[1].get('uso', ''))
                ws_r2.cell(row=row, column=29, value=zonas[1].get('puntaje', 0))
                ws_r2.cell(row=row, column=30, value=zonas[1].get('area_construida', 0))
            
            # Construcción 3 (columnas 31-38)
            if len(zonas) >= 3:
                ws_r2.cell(row=row, column=31, value=zonas[2].get('habitaciones', 0))
                ws_r2.cell(row=row, column=32, value=zonas[2].get('banos', 0))
                ws_r2.cell(row=row, column=33, value=zonas[2].get('locales', 0))
                ws_r2.cell(row=row, column=34, value=zonas[2].get('pisos', 0))
                ws_r2.cell(row=row, column=35, value=zonas[2].get('tipificacion', ''))
                ws_r2.cell(row=row, column=36, value=zonas[2].get('uso', ''))
                ws_r2.cell(row=row, column=37, value=zonas[2].get('puntaje', 0))
                ws_r2.cell(row=row, column=38, value=zonas[2].get('area_construida', 0))
            
            # Vigencia
            ws_r2.cell(row=row, column=39, value=predio.get('vigencia', datetime.now().year))
            row += 1
    
    # Ajustar anchos de columna
    for ws in [ws_r1, ws_r2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generar nombre de archivo
    fecha = datetime.now().strftime('%Y%m%d')
    filename = f"Predios_{municipio or 'Todos'}_{fecha}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ===== CERTIFICADO CATASTRAL =====

def generate_certificado_catastral(predio: dict, firmante: dict, proyectado_por: str, numero_certificado: str = None) -> bytes:
    """
    Genera un certificado catastral en PDF siguiendo el diseño institucional de Asomunicipios.
    Formato de número: COM-F03-XXXX-GC-XX
    Basado en el diseño de CERTIFICADO_CATASTRAL.pdf
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import simpleSplit
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Colores EXACTOS según el PDF de referencia
    azul_titulo = colors.HexColor('#1e3a5f')  # Azul oscuro para títulos
    azul_celeste = colors.HexColor('#CCE5FF')  # Azul celeste pálido para encabezado sección
    verde_seccion = colors.HexColor('#2d7a4f')  # Verde para subsecciones
    gris_texto = colors.HexColor('#333333')
    gris_claro = colors.HexColor('#666666')
    linea_gris = colors.HexColor('#dddddd')
    
    # Márgenes
    left_margin = 1.5 * cm
    right_margin = width - 1.5 * cm
    content_width = right_margin - left_margin
    
    fecha_actual = datetime.now()
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    
    # === ENCABEZADO ===
    y = height - 1 * cm
    
    # Logo de Asomunicipios (izquierda) - más grande como en el original
    logo_path = Path("/app/backend/logo_asomunicipios.jpeg")
    if not logo_path.exists():
        logo_path = Path("/app/backend/logo_asomunicipios.png")
    if logo_path.exists():
        logo_width = 5 * cm
        logo_height = 2.8 * cm
        c.drawImage(str(logo_path), left_margin, height - 3.8 * cm, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')
    
    # Texto encabezado (centro-izquierda, a la derecha del logo)
    header_x = left_margin + 5.5 * cm
    c.setFillColor(azul_titulo)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(header_x, y, "Asomunicipios")
    y -= 12
    c.setFont("Helvetica", 8)
    c.setFillColor(gris_texto)
    c.drawString(header_x, y, "Asociación de Municipios del Catatumbo")
    y -= 10
    c.drawString(header_x, y, "Provincia de Ocaña y Sur del Cesar")
    y -= 12
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(verde_seccion)
    c.drawString(header_x, y, "Gestor Catastral")
    
    # Fecha (derecha superior)
    fecha_str = f"{fecha_actual.day} de {meses[fecha_actual.month-1]} del {fecha_actual.year}"
    c.setFont("Helvetica", 9)
    c.setFillColor(gris_texto)
    c.drawRightString(right_margin, height - 1 * cm, fecha_str)
    
    # Número de certificado (derecha, debajo de fecha) - COM-F03-XXX-GC-XX
    cert_numero = numero_certificado or "COM-F03-____-GC-__"
    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(right_margin, height - 1.5 * cm, f"CERTIFICADO: {cert_numero}")
    
    y = height - 4 * cm
    
    # Línea separadora verde
    c.setStrokeColor(verde_seccion)
    c.setLineWidth(1.5)
    c.line(left_margin, y, right_margin, y)
    y -= 20
    
    # === TÍTULO PRINCIPAL ===
    c.setFillColor(azul_titulo)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width/2, y, "CERTIFICADO CATASTRAL")
    y -= 14
    
    # Base legal (texto pequeño centrado)
    c.setFillColor(gris_claro)
    c.setFont("Helvetica", 6)
    texto_legal1 = "LEY 527 DE 1999 (AGOSTO 18), Directiva Presidencial No. 02 del 2000, Ley 962 de 2005 (Anti trámites), Articulo 6, Parágrafo 3."
    texto_legal2 = "artículo 2.2.2.2.8 del Decreto 148 de 2020, artículo 29 de la resolución No. 1149 de 2021 emanada del Instituto Geográfico Agustín Codazzi"
    c.drawCentredString(width/2, y, texto_legal1)
    y -= 9
    c.drawCentredString(width/2, y, texto_legal2)
    y -= 16
    
    # === TEXTO CERTIFICADOR ===
    c.setFillColor(gris_texto)
    c.setFont("Helvetica", 9)
    intro = "LA ASOCIACIÓN DE MUNICIPIOS DEL CATATUMBO PROVINCIA DE OCAÑA Y SUR DEL CESAR – ASOMUNICIPIOS, en su calidad de Gestor Catastral habilitado mediante resolución No. 1092 del 18 de diciembre de 2020, del Instituto Geográfico Agustín Codazzi – IGAC,"
    lines = simpleSplit(intro, "Helvetica", 9, content_width)
    for line in lines:
        c.drawString(left_margin, y, line)
        y -= 11
    y -= 4
    
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(width/2, y, "CERTIFICA:")
    y -= 14
    
    c.setFont("Helvetica", 9)
    certifica_texto = "Que en la base de datos catastral a cargo de esta entidad se encuentra inscrito el siguiente predio:"
    c.drawString(left_margin, y, certifica_texto)
    y -= 18
    
    # === ENCABEZADO SECCIÓN - AZUL CELESTE PÁLIDO (como en el original) ===
    c.setFillColor(azul_celeste)
    c.rect(left_margin, y - 14, content_width, 17, fill=1, stroke=0)
    c.setFillColor(azul_titulo)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(width/2, y - 10, "INFORMACIÓN CATASTRAL DEL PREDIO")
    y -= 22
    
    # Predio No. (alineado a la derecha)
    c.setFillColor(gris_texto)
    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(right_margin - 5, y, "Predio No. 01")
    y -= 14
    
    # Función para dibujar fila de campo (SOLO líneas horizontales como en el original)
    def draw_field(label, value, y_pos):
        c.setFillColor(gris_texto)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(left_margin + 3, y_pos, label)
        c.setFont("Helvetica", 8)
        c.drawString(left_margin + 130, y_pos, str(value) if value else "")
        # Solo línea horizontal separadora (como en el original)
        c.setStrokeColor(linea_gris)
        c.setLineWidth(0.5)
        c.line(left_margin, y_pos - 4, right_margin, y_pos - 4)
        return y_pos - 13
    
    # === INFORMACIÓN JURÍDICA (con fondo celeste en encabezado) ===
    c.setFillColor(azul_celeste)
    c.rect(left_margin, y - 12, content_width, 15, fill=1, stroke=0)
    c.setFillColor(azul_titulo)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left_margin + 5, y - 8, "INFORMACIÓN JURÍDICA")
    y -= 20
    
    propietarios = predio.get('propietarios', [])
    if propietarios:
        prop = propietarios[0]
        y = draw_field("Nombre de los propietarios:", prop.get('nombre_propietario', ''), y)
        y = draw_field("Número de propietario:", "01", y)
        y = draw_field("Tipo de documento:", prop.get('tipo_documento', 'C'), y)
        y = draw_field("Número de documento:", prop.get('numero_documento', ''), y)
    else:
        y = draw_field("Nombre de los propietarios:", predio.get('nombre_propietario', 'N/A'), y)
        y = draw_field("Número de propietario:", "01", y)
        y = draw_field("Tipo de documento:", predio.get('tipo_documento', 'C'), y)
        y = draw_field("Número de documento:", predio.get('numero_documento', ''), y)
    
    matricula = ''
    r2_registros = predio.get('r2_registros', [])
    if r2_registros:
        matricula = r2_registros[0].get('matricula_inmobiliaria', '')
    y = draw_field("Matrícula:", matricula or 'N/A', y)
    y -= 8
    
    # === INFORMACIÓN FÍSICA (con fondo celeste en encabezado) ===
    c.setFillColor(azul_celeste)
    c.rect(left_margin, y - 12, content_width, 15, fill=1, stroke=0)
    c.setFillColor(azul_titulo)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left_margin + 5, y - 8, "INFORMACIÓN FÍSICA")
    y -= 20
    
    municipio = predio.get('municipio', '')
    if municipio in ['Río de Oro']:
        depto_cod = "20 - CESAR"
        muni_cod = "614 - RIO DE ORO"
    else:
        depto_cod = "54 - NORTE DE SANTANDER"
        muni_mapping = {
            'Ábrego': '003 - ÁBREGO', 'Bucarasica': '109 - BUCARASICA',
            'Convención': '206 - CONVENCIÓN', 'Cáchira': '128 - CÁCHIRA',
            'El Carmen': '245 - EL CARMEN', 'El Tarra': '250 - EL TARRA',
            'Hacarí': '344 - HACARÍ', 'La Playa': '398 - LA PLAYA',
            'San Calixto': '670 - SAN CALIXTO', 'Sardinata': '720 - SARDINATA',
            'Teorama': '800 - TEORAMA'
        }
        muni_cod = muni_mapping.get(municipio, municipio)
    
    y = draw_field("Departamento:", depto_cod, y)
    y = draw_field("Municipio:", muni_cod, y)
    y = draw_field("Número predial:", predio.get('codigo_predial_nacional', ''), y)
    y = draw_field("Número predial anterior:", predio.get('codigo_homologado', ''), y)
    y = draw_field("Dirección:", predio.get('direccion', ''), y)
    
    area_terreno = predio.get('area_terreno', 0)
    y = draw_field("Área terreno:", f"{int(area_terreno):,} m²".replace(',', '.'), y)
    area_construida = predio.get('area_construida', 0)
    y = draw_field("Área construida:", f"{int(area_construida):,} m²".replace(',', '.'), y)
    y -= 8
    
    # === INFORMACIÓN ECONÓMICA (con fondo celeste en encabezado) ===
    c.setFillColor(azul_celeste)
    c.rect(left_margin, y - 12, content_width, 15, fill=1, stroke=0)
    c.setFillColor(azul_titulo)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left_margin + 5, y - 8, "INFORMACIÓN ECONÓMICA")
    y -= 20
    
    avaluo = predio.get('avaluo', 0)
    avaluo_str = f"${int(avaluo):,}".replace(',', '.')
    y = draw_field("Avalúo catastral:", avaluo_str, y)
    y -= 14
    
    # === TEXTO DE EXPEDICIÓN ===
    c.setFillColor(gris_texto)
    c.setFont("Helvetica", 9)
    fecha_str = f"{fecha_actual.day} de {meses[fecha_actual.month-1]} del {fecha_actual.year}"
    texto_exp = f"El presente certificado se expide a favor del interesado el {fecha_str}."
    c.drawString(left_margin, y, texto_exp)
    y -= 30
    
    # === FIRMAS ===
    c.setFont("Helvetica", 8)
    c.setFillColor(gris_texto)
    c.drawString(left_margin, y, f"Elaboró: {proyectado_por}")
    c.drawString(left_margin + 180, y, "Revisó: Juan C. Alsina")
    y -= 28
    
    # Firma principal
    firma_x = width/2
    c.setStrokeColor(gris_texto)
    c.setLineWidth(0.5)
    c.line(firma_x - 75, y + 12, firma_x + 75, y + 12)
    
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(azul_titulo)
    c.drawCentredString(firma_x, y, "DALGIE ESPERANZA TORRADO RIZO")
    y -= 11
    c.setFont("Helvetica", 8)
    c.setFillColor(gris_texto)
    c.drawCentredString(firma_x, y, "SUBDIRECTORA FINANCIERA Y ADMINISTRATIVA")
    y -= 24
    
    # === NOTAS ===
    c.setFillColor(gris_claro)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(left_margin, y, "NOTA:")
    y -= 9
    
    c.setFont("Helvetica", 6)
    notas = [
        "• La presente información no sirve como prueba para establecer actos constitutivos de posesión.",
        "• De conformidad con el artículo 2.2.2.2.8 del Decreto 148 de 2020, Inscripción o incorporación catastral. La información catastral resultado de los procesos de formación,",
        "  actualización o conservación se inscribirá o incorporará en la base catastral con la fecha del acto administrativo que lo ordena.",
        "• Adicionalmente de conformidad con el artículo 29 de la resolución No. 1149 de 2021 emanada del Instituto Geográfico Agustín Codazzi, \"Efecto jurídico de la inscripción",
        "  catastral. La inscripción en el catastro no constituye título de dominio, ni sanea los vicios de que adolezca la titulación presentada o la posesión del interesado.\"",
        "• La base catastral de Asomunicipios sólo incluye información de los municipios habilitados dentro del esquema asociativo.",
        "• Ante cualquier inquietud, puede escribir al correo electrónico: comunicaciones@asomunicipios.gov.co",
    ]
    
    for nota in notas:
        c.drawString(left_margin, y, nota)
        y -= 8
    
    y -= 10
    
    # === PIE DE PÁGINA ===
    c.setStrokeColor(verde_seccion)
    c.setLineWidth(1)
    c.line(left_margin, y, right_margin, y)
    y -= 12
    
    c.setFillColor(gris_texto)
    c.setFont("Helvetica", 7)
    # Redes sociales e info de contacto (como en el original)
    c.drawString(left_margin, y, "f @asomunicipios    X @asomunicipios    IG @asomunicipios")
    c.drawCentredString(width/2, y, "comunicaciones@asomunicipios.gov.co")
    c.drawRightString(right_margin, y, "Calle 12 # 11-76 Ocaña | +57 3102327647")
    
    c.save()
    return buffer.getvalue()


@api_router.get("/predios/{predio_id}/certificado")
async def generar_certificado_catastral_endpoint(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Genera un certificado catastral PDF para un predio específico con consecutivo en blanco para llenado manual"""
    # Solo coordinador, administrador y atencion_usuario pueden generar certificados
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR, UserRole.ATENCION_USUARIO]:
        raise HTTPException(status_code=403, detail="No tiene permiso para generar certificados")
    
    # Obtener predio
    predio = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Registrar certificado en la base de datos (sin número, se llena manualmente)
    certificado_record = {
        "id": str(uuid.uuid4()),
        "numero": "(Por asignar)",
        "predio_id": predio_id,
        "codigo_predial": predio.get('codigo_predial_nacional', ''),
        "generado_por": current_user['id'],
        "generado_por_nombre": current_user['full_name'],
        "generado_por_rol": current_user['role'],
        "fecha_generacion": datetime.now(timezone.utc).isoformat()
    }
    await db.certificados.insert_one(certificado_record)
    
    # Firmante siempre es Dalgie Esperanza Torrado Rizo
    firmante = {
        "full_name": "DALGIE ESPERANZA TORRADO RIZO",
        "cargo": "Subdirectora Financiera y Administrativa"
    }
    
    # Quien proyecta es el usuario actual
    proyectado_por = current_user['full_name']
    
    # Generar PDF con campo editable para número
    pdf_bytes = generate_certificado_catastral(predio, firmante, proyectado_por)
    
    # Guardar temporalmente
    temp_path = UPLOAD_DIR / f"certificado_{predio_id}_{uuid.uuid4()}.pdf"
    with open(temp_path, 'wb') as f:
        f.write(pdf_bytes)
    
    # Nombre del archivo
    codigo = predio.get('codigo_predial_nacional', predio_id)
    filename = f"Certificado_Catastral_{codigo}.pdf"
    
    return FileResponse(
        path=temp_path,
        filename=filename,
        media_type='application/pdf'
    )


@api_router.get("/certificados/historial")
async def get_certificados_historial(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el historial de certificados generados"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    certificados = await db.certificados.find({}, {"_id": 0}).sort("fecha_generacion", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.certificados.count_documents({})
    
    return {
        "certificados": certificados,
        "total": total
    }


@api_router.get("/predios/terreno-info/{municipio}")
async def get_terreno_info(
    municipio: str,
    zona: str = "00",
    sector: str = "01", 
    manzana_vereda: str = "0000",
    current_user: dict = Depends(get_current_user)
):
    """Obtiene información sobre el siguiente terreno disponible en una manzana"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Buscar todos los terrenos en esta manzana (incluyendo eliminados)
    query = {
        "municipio": municipio,
        "zona": zona,
        "sector": sector,
        "manzana_vereda": manzana_vereda
    }
    
    predios = await db.predios.find(query, {"_id": 0, "terreno": 1, "terreno_num": 1, "deleted": 1, "codigo_homologado": 1}).to_list(10000)
    
    # Clasificar terrenos
    terrenos_activos = []
    terrenos_eliminados = []
    
    for p in predios:
        terreno_num = p.get('terreno_num', 0)
        if p.get('deleted'):
            terrenos_eliminados.append({
                "numero": p.get('terreno'),
                "codigo": p.get('codigo_homologado')
            })
        else:
            terrenos_activos.append(terreno_num)
    
    # Encontrar el máximo terreno usado
    max_terreno = max(terrenos_activos + [t.get('terreno_num', 0) for t in predios], default=0)
    siguiente_terreno = max_terreno + 1
    
    return {
        "municipio": municipio,
        "zona": zona,
        "sector": sector,
        "manzana_vereda": manzana_vereda,
        "total_activos": len(terrenos_activos),
        "ultimo_terreno": str(max_terreno).zfill(4) if max_terreno > 0 else "N/A",
        "siguiente_terreno": str(siguiente_terreno).zfill(4),
        "terrenos_eliminados": terrenos_eliminados,
        "terrenos_no_reutilizables": len(terrenos_eliminados)
    }

@api_router.get("/predios/{predio_id}")
async def get_predio(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene un predio por ID"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    return predio

@api_router.post("/predios")
async def create_predio(predio_data: PredioCreate, current_user: dict = Depends(get_current_user)):
    """Crea un nuevo predio"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    r1 = predio_data.r1
    r2 = predio_data.r2
    
    # Obtener siguiente número de terreno
    terreno, terreno_num = await get_next_terreno_number(
        r1.municipio, r1.zona, r1.sector, r1.manzana_vereda
    )
    
    # Generar código predial nacional
    codigo_predial = await generate_codigo_predial(
        r1.municipio, r1.zona, r1.sector, r1.manzana_vereda,
        terreno, r1.condicion_predio, r1.predio_horizontal
    )
    
    # Verificar que no exista (incluyendo eliminados)
    existing = await db.predios.find_one({"codigo_predial_nacional": codigo_predial})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un predio con este código predial")
    
    # Generar código homologado
    codigo_homologado, numero_predio = await generate_codigo_homologado(r1.municipio)
    
    # Obtener código Código Nacional Catastral
    divipola = MUNICIPIOS_DIVIPOLA[r1.municipio]
    
    # Crear el predio
    predio = {
        "id": str(uuid.uuid4()),
        "departamento": divipola["departamento"],
        "municipio": r1.municipio,
        "municipio_codigo": divipola["municipio"],
        "numero_predio": numero_predio,
        "codigo_predial_nacional": codigo_predial,
        "codigo_homologado": codigo_homologado,
        "zona": r1.zona,
        "sector": r1.sector,
        "manzana_vereda": r1.manzana_vereda,
        "terreno": terreno,
        "terreno_num": terreno_num,
        "condicion_predio": r1.condicion_predio,
        "predio_horizontal": r1.predio_horizontal,
        
        # R1 - Información jurídica
        "tipo_registro": 1,
        "nombre_propietario": r1.nombre_propietario,
        "tipo_documento": r1.tipo_documento,
        "numero_documento": r1.numero_documento,
        "estado_civil": r1.estado_civil,
        "direccion": r1.direccion,
        "comuna": r1.comuna,
        "destino_economico": r1.destino_economico,
        "area_terreno": r1.area_terreno,
        "area_construida": r1.area_construida,
        "avaluo": r1.avaluo,
        "tipo_mutacion": r1.tipo_mutacion,
        "numero_resolucion": r1.numero_resolucion,
        "fecha_resolucion": r1.fecha_resolucion,
        
        # R2 - Información física (si se proporciona)
        "r2": r2.model_dump() if r2 else None,
        
        # Metadata
        "vigencia": datetime.now().strftime("%m%d%Y"),
        "created_by": current_user['id'],
        "created_by_name": current_user['full_name'],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted": False,
        
        # Historial
        "historial": [{
            "accion": "Predio creado",
            "usuario": current_user['full_name'],
            "usuario_id": current_user['id'],
            "fecha": datetime.now(timezone.utc).isoformat()
        }]
    }
    
    await db.predios.insert_one(predio)
    
    # Remover _id antes de retornar
    predio.pop("_id", None)
    
    return predio

@api_router.patch("/predios/{predio_id}")
async def update_predio(predio_id: str, update_data: PredioUpdate, current_user: dict = Depends(get_current_user)):
    """Actualiza un predio existente"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Filtrar campos no nulos
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if not update_dict:
        return predio
    
    # Agregar metadata de actualización
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Agregar al historial
    historial_entry = {
        "accion": "Predio modificado",
        "usuario": current_user['full_name'],
        "usuario_id": current_user['id'],
        "campos_modificados": list(update_dict.keys()),
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios.update_one(
        {"id": predio_id},
        {
            "$set": update_dict,
            "$push": {"historial": historial_entry}
        }
    )
    
    # Retornar predio actualizado
    updated_predio = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    return updated_predio

@api_router.delete("/predios/{predio_id}")
async def delete_predio(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Elimina un predio (soft delete)"""
    if current_user['role'] not in [UserRole.ADMINISTRADOR, UserRole.COORDINADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores y administradores pueden eliminar predios")
    
    predio = await db.predios.find_one({"id": predio_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    # Soft delete - NO eliminamos físicamente para evitar reutilizar códigos
    historial_entry = {
        "accion": "Predio eliminado",
        "usuario": current_user['full_name'],
        "usuario_id": current_user['id'],
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    
    await db.predios.update_one(
        {"id": predio_id},
        {
            "$set": {
                "deleted": True,
                "deleted_at": datetime.now(timezone.utc).isoformat(),
                "deleted_by": current_user['id'],
                "deleted_by_name": current_user['full_name']
            },
            "$push": {"historial": historial_entry}
        }
    )
    
    return {"message": "Predio eliminado exitosamente"}


# ===== SISTEMA DE APROBACIÓN DE PREDIOS =====

@api_router.post("/predios/cambios/proponer")
async def proponer_cambio_predio(
    cambio: CambioPendienteCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Propone un cambio en un predio (crear, modificar, eliminar).
    Solo gestores y atención pueden proponer. Coordinadores aprueban directamente.
    """
    # Verificar permisos
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Coordinadores y administradores aprueban directamente
    aprueba_directo = current_user['role'] in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]
    
    cambio_doc = {
        "id": str(uuid.uuid4()),
        "predio_id": cambio.predio_id,
        "tipo_cambio": cambio.tipo_cambio,
        "datos_propuestos": cambio.datos_propuestos,
        "justificacion": cambio.justificacion,
        "estado": PredioEstadoAprobacion.APROBADO if aprueba_directo else f"pendiente_{cambio.tipo_cambio}",
        "propuesto_por": current_user['id'],
        "propuesto_por_nombre": current_user['full_name'],
        "propuesto_por_rol": current_user['role'],
        "fecha_propuesta": datetime.now(timezone.utc).isoformat(),
        "aprobado_por": current_user['id'] if aprueba_directo else None,
        "aprobado_por_nombre": current_user['full_name'] if aprueba_directo else None,
        "fecha_aprobacion": datetime.now(timezone.utc).isoformat() if aprueba_directo else None,
        "comentario_aprobacion": "Aprobación directa por coordinador/administrador" if aprueba_directo else None
    }
    
    # Si aprueba directo, aplicar el cambio inmediatamente
    if aprueba_directo:
        resultado = await aplicar_cambio_predio(cambio_doc, current_user)
        cambio_doc["resultado"] = resultado
    
    # Guardar el cambio en la colección de cambios
    await db.predios_cambios.insert_one(cambio_doc)
    
    return {
        "id": cambio_doc["id"],
        "estado": cambio_doc["estado"],
        "mensaje": "Cambio aplicado directamente" if aprueba_directo else "Cambio propuesto, pendiente de aprobación",
        "requiere_aprobacion": not aprueba_directo
    }


@api_router.get("/predios/cambios/pendientes")
async def get_cambios_pendientes(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista todos los cambios pendientes de aprobación (solo coordinadores/admin)"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden ver cambios pendientes")
    
    query = {
        "estado": {"$in": [
            PredioEstadoAprobacion.PENDIENTE_CREACION,
            PredioEstadoAprobacion.PENDIENTE_MODIFICACION,
            PredioEstadoAprobacion.PENDIENTE_ELIMINACION
        ]}
    }
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_propuesta", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enriquecer con datos del predio actual si existe
    for cambio in cambios:
        if cambio.get("predio_id"):
            predio = await db.predios.find_one({"id": cambio["predio_id"]}, {"_id": 0, "codigo_homologado": 1, "nombre_propietario": 1, "municipio": 1})
            cambio["predio_actual"] = predio
    
    return {
        "total": total,
        "cambios": cambios
    }


@api_router.get("/predios/cambios/historial")
async def get_historial_cambios(
    predio_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene el historial de cambios (aprobados y rechazados)"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {}
    if predio_id:
        query["predio_id"] = predio_id
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_propuesta", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "cambios": cambios
    }


@api_router.post("/predios/cambios/aprobar")
async def aprobar_rechazar_cambio(
    request: CambioAprobacionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Aprueba o rechaza un cambio pendiente (solo coordinadores/admin)"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden aprobar cambios")
    
    # Buscar el cambio
    cambio = await db.predios_cambios.find_one({"id": request.cambio_id}, {"_id": 0})
    
    if not cambio:
        raise HTTPException(status_code=404, detail="Cambio no encontrado")
    
    if cambio["estado"] not in [
        PredioEstadoAprobacion.PENDIENTE_CREACION,
        PredioEstadoAprobacion.PENDIENTE_MODIFICACION,
        PredioEstadoAprobacion.PENDIENTE_ELIMINACION
    ]:
        raise HTTPException(status_code=400, detail="Este cambio ya fue procesado")
    
    nuevo_estado = PredioEstadoAprobacion.APROBADO if request.aprobado else PredioEstadoAprobacion.RECHAZADO
    
    update_data = {
        "estado": nuevo_estado,
        "aprobado_por": current_user['id'],
        "aprobado_por_nombre": current_user['full_name'],
        "fecha_aprobacion": datetime.now(timezone.utc).isoformat(),
        "comentario_aprobacion": request.comentario
    }
    
    # Si se aprueba, aplicar el cambio
    if request.aprobado:
        resultado = await aplicar_cambio_predio(cambio, current_user)
        update_data["resultado"] = resultado
    
    await db.predios_cambios.update_one(
        {"id": request.cambio_id},
        {"$set": update_data}
    )
    
    return {
        "mensaje": "Cambio aprobado y aplicado" if request.aprobado else "Cambio rechazado",
        "estado": nuevo_estado
    }


async def aplicar_cambio_predio(cambio: dict, aprobador: dict) -> dict:
    """Aplica un cambio aprobado al predio"""
    tipo = cambio["tipo_cambio"]
    datos = cambio["datos_propuestos"]
    
    historial_entry = {
        "accion": f"Cambio {tipo} aprobado",
        "usuario": aprobador['full_name'],
        "usuario_rol": aprobador['role'],
        "propuesto_por": cambio.get("propuesto_por_nombre"),
        "fecha": datetime.now(timezone.utc).isoformat(),
        "comentario": cambio.get("comentario_aprobacion")
    }
    
    if tipo == "creacion":
        # Crear nuevo predio
        predio_doc = datos.copy()
        predio_doc["id"] = str(uuid.uuid4())
        predio_doc["estado_aprobacion"] = PredioEstadoAprobacion.APROBADO
        predio_doc["deleted"] = False
        predio_doc["created_at"] = datetime.now(timezone.utc).isoformat()
        predio_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
        predio_doc["historial"] = [historial_entry]
        
        await db.predios.insert_one(predio_doc)
        return {"predio_id": predio_doc["id"], "accion": "creado"}
    
    elif tipo == "modificacion":
        predio_id = cambio["predio_id"]
        
        # Actualizar predio
        datos["updated_at"] = datetime.now(timezone.utc).isoformat()
        datos["estado_aprobacion"] = PredioEstadoAprobacion.APROBADO
        
        await db.predios.update_one(
            {"id": predio_id},
            {
                "$set": datos,
                "$push": {"historial": historial_entry}
            }
        )
        return {"predio_id": predio_id, "accion": "modificado"}
    
    elif tipo == "eliminacion":
        predio_id = cambio["predio_id"]
        
        # Soft delete
        await db.predios.update_one(
            {"id": predio_id},
            {
                "$set": {
                    "deleted": True,
                    "deleted_at": datetime.now(timezone.utc).isoformat(),
                    "deleted_by": aprobador['id'],
                    "deleted_by_name": aprobador['full_name'],
                    "estado_aprobacion": PredioEstadoAprobacion.APROBADO
                },
                "$push": {"historial": historial_entry}
            }
        )
        return {"predio_id": predio_id, "accion": "eliminado"}
    
    return {"error": "Tipo de cambio no reconocido"}


@api_router.get("/predios/cambios/mis-propuestas")
async def get_mis_propuestas(
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista las propuestas de cambio del usuario actual"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"propuesto_por": current_user['id']}
    
    total = await db.predios_cambios.count_documents(query)
    cambios = await db.predios_cambios.find(query, {"_id": 0}).sort("fecha_propuesta", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "cambios": cambios
    }


@api_router.get("/predios/cambios/stats")
async def get_cambios_stats(current_user: dict = Depends(get_current_user)):
    """Obtiene estadísticas de cambios pendientes"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    pendientes_creacion = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.PENDIENTE_CREACION})
    pendientes_modificacion = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.PENDIENTE_MODIFICACION})
    pendientes_eliminacion = await db.predios_cambios.count_documents({"estado": PredioEstadoAprobacion.PENDIENTE_ELIMINACION})
    
    return {
        "pendientes_creacion": pendientes_creacion,
        "pendientes_modificacion": pendientes_modificacion,
        "pendientes_eliminacion": pendientes_eliminacion,
        "total_pendientes": pendientes_creacion + pendientes_modificacion + pendientes_eliminacion
    }


# ===== GEOGRAPHIC DATABASE (GDB) INTEGRATION =====

GDB_PATH = Path("/app/gdb_data/54003.gdb")

def get_gdb_geometry(codigo_predial: str) -> Optional[dict]:
    """Get geometry for a property from multiple GDB files, transformed to WGS84 for web mapping"""
    import geopandas as gpd
    from shapely.geometry import mapping
    
    # Mapeo de códigos de municipio a archivos GDB
    GDB_FILES = {
        '54003': '/app/gdb_data/54003.gdb',  # Ábrego
        '54109': '/app/gdb_data/54109.gdb',  # Bucarasica
        '54128': '/app/gdb_data/54128.gdb',  # Cáchira
    }
    
    try:
        # Extraer código de municipio del código predial (posiciones 0-5)
        municipio_code = codigo_predial[:5]
        
        # Determinar si es rural o urbano
        sector = codigo_predial[5:8]
        is_urban = sector != "000"
        
        # Buscar el archivo GDB correcto
        gdb_path = None
        for code, path in GDB_FILES.items():
            if municipio_code == code:
                gdb_path = Path(path)
                break
        
        if not gdb_path or not gdb_path.exists():
            # Intentar con el GDB por defecto
            if GDB_PATH.exists():
                gdb_path = GDB_PATH
            else:
                return None
        
        # Determinar nombre de capa (diferentes GDBs usan diferentes nombres)
        # 54003 (Ábrego): R_TERRENO_1, U_TERRENO_1
        # Otros: R_TERRENO, U_TERRENO
        if '54003' in str(gdb_path):
            layer = "U_TERRENO_1" if is_urban else "R_TERRENO_1"
        else:
            layer = "U_TERRENO" if is_urban else "R_TERRENO"
        
        # Leer capa y buscar el código
        gdf = gpd.read_file(str(gdb_path), layer=layer)
        match = gdf[gdf['codigo'] == codigo_predial]
        
        if len(match) == 0:
            return None
        
        # Obtener área y perímetro originales
        area_m2 = float(match.iloc[0]['shape_Area']) if 'shape_Area' in match.columns else None
        perimetro_m = float(match.iloc[0]['shape_Length']) if 'shape_Length' in match.columns else None
        
        # Transformar a WGS84
        match_wgs84 = match.to_crs(epsg=4326)
        
        geom = match_wgs84.iloc[0]['geometry']
        if geom is None:
            return None
            
        geojson = mapping(geom)
        
        return {
            "type": "Feature",
            "geometry": geojson,
            "properties": {
                "codigo": codigo_predial,
                "area_m2": area_m2,
                "perimetro_m": perimetro_m,
                "tipo": "Urbano" if is_urban else "Rural"
            }
        }
    except Exception as e:
        logger.error(f"Error reading GDB geometry: {e}")
        return None


@api_router.get("/predios/{predio_id}/geometria")
async def get_predio_geometry(predio_id: str, current_user: dict = Depends(get_current_user)):
    """Get geographic geometry for a property"""
    # Only staff can access geometry
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    # Get predio from database
    predio = await db.predios.find_one({"id": predio_id}, {"_id": 0})
    if not predio:
        raise HTTPException(status_code=404, detail="Predio no encontrado")
    
    codigo = predio.get("codigo_predial_nacional")
    if not codigo:
        raise HTTPException(status_code=404, detail="Predio sin código catastral")
    
    geometry = get_gdb_geometry(codigo)
    if not geometry:
        raise HTTPException(status_code=404, detail="Geometría no disponible para este predio")
    
    return geometry


@api_router.get("/predios/codigo/{codigo_predial}/geometria")
async def get_geometry_by_code(codigo_predial: str, current_user: dict = Depends(get_current_user)):
    """Get geographic geometry directly by cadastral code"""
    # Only staff can access geometry
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    geometry = get_gdb_geometry(codigo_predial)
    if not geometry:
        raise HTTPException(status_code=404, detail="Geometría no disponible para este código")
    
    return geometry


@api_router.get("/gdb/geometrias")
async def get_geometrias_filtradas(
    municipio: Optional[str] = None,
    zona: Optional[str] = None,  # 'urbano' o 'rural'
    limit: int = 500,
    current_user: dict = Depends(get_current_user)
):
    """Get all geometries for a municipality/zone to display on map"""
    import geopandas as gpd
    from pyproj import CRS, Transformer
    from shapely.ops import transform
    
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    if not municipio:
        raise HTTPException(status_code=400, detail="Debe especificar un municipio")
    
    # GDB files configuration
    GDB_FILES = {
        'Ábrego': ('/app/gdb_data/54003.gdb', 'R_TERRENO_1', 'U_TERRENO_1'),
        'Bucarasica': ('/app/gdb_data/54109.gdb', 'R_TERRENO', 'U_TERRENO'),
        'Cáchira': ('/app/gdb_data/54128.gdb', 'R_TERRENO', 'U_TERRENO'),
    }
    
    if municipio not in GDB_FILES:
        raise HTTPException(status_code=404, detail=f"No hay datos geográficos para {municipio}")
    
    gdb_path, rural_layer, urban_layer = GDB_FILES[municipio]
    
    if not Path(gdb_path).exists():
        raise HTTPException(status_code=404, detail=f"Archivo GDB no encontrado para {municipio}")
    
    try:
        features = []
        
        # Set up coordinate transformer (EPSG:3116 to WGS84)
        transformer = Transformer.from_crs(CRS("EPSG:3116"), CRS("EPSG:4326"), always_xy=True)
        
        def transform_coords(geom):
            return transform(transformer.transform, geom)
        
        # Read requested zone(s)
        layers_to_read = []
        if zona == 'urbano':
            layers_to_read = [(urban_layer, 'Urbano')]
        elif zona == 'rural':
            layers_to_read = [(rural_layer, 'Rural')]
        else:
            layers_to_read = [(rural_layer, 'Rural'), (urban_layer, 'Urbano')]
        
        for layer_name, tipo in layers_to_read:
            try:
                gdf = gpd.read_file(gdb_path, layer=layer_name)
                
                # Find the code column
                code_col = None
                for col in ['codigo_ant', 'CODIGO', 'codigo', 'CODIGO_ANT']:
                    if col in gdf.columns:
                        code_col = col
                        break
                
                # Limit results
                gdf = gdf.head(limit)
                
                for idx, row in gdf.iterrows():
                    geom = row.geometry
                    if geom is not None:
                        # Transform coordinates
                        geom_wgs84 = transform_coords(geom)
                        
                        codigo = str(row[code_col]) if code_col and row[code_col] else f"SIN_CODIGO_{idx}"
                        
                        feature = {
                            "type": "Feature",
                            "geometry": geom_wgs84.__geo_interface__,
                            "properties": {
                                "codigo": codigo,
                                "tipo": tipo,
                                "area_m2": round(geom.area, 2) if geom.area else 0
                            }
                        }
                        features.append(feature)
            except Exception as e:
                logger.warning(f"Error reading layer {layer_name}: {e}")
                continue
        
        return {
            "type": "FeatureCollection",
            "municipio": municipio,
            "zona_filter": zona,
            "total": len(features),
            "features": features
        }
    except Exception as e:
        logger.error(f"Error getting geometries: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener geometrías: {str(e)}")


@api_router.get("/gdb/stats")
async def get_gdb_stats(current_user: dict = Depends(get_current_user)):
    """Get statistics about available geographic data from all GDB files"""
    import geopandas as gpd
    
    # Only staff can access stats
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    GDB_FILES = {
        'Ábrego': ('/app/gdb_data/54003.gdb', 'R_TERRENO_1', 'U_TERRENO_1'),
        'Bucarasica': ('/app/gdb_data/54109.gdb', 'R_TERRENO', 'U_TERRENO'),
        'Cáchira': ('/app/gdb_data/54128.gdb', 'R_TERRENO', 'U_TERRENO'),
    }
    
    try:
        stats_by_municipio = {}
        total_rurales = 0
        total_urbanos = 0
        
        for municipio, (path, rural_layer, urban_layer) in GDB_FILES.items():
            if Path(path).exists():
                try:
                    r_count = len(gpd.read_file(path, layer=rural_layer))
                    u_count = len(gpd.read_file(path, layer=urban_layer))
                    stats_by_municipio[municipio] = {
                        "rurales": r_count,
                        "urbanos": u_count,
                        "total": r_count + u_count
                    }
                    total_rurales += r_count
                    total_urbanos += u_count
                except Exception as e:
                    stats_by_municipio[municipio] = {"error": str(e)}
        
        return {
            "gdb_disponible": True,
            "predios_rurales": total_rurales,
            "predios_urbanos": total_urbanos,
            "total_geometrias": total_rurales + total_urbanos,
            "municipios": stats_by_municipio
        }
    except Exception as e:
        logger.error(f"Error reading GDB stats: {e}")
        raise HTTPException(status_code=500, detail=f"Error leyendo base de datos geográfica: {str(e)}")


@api_router.get("/gdb/capas")
async def get_gdb_layers(current_user: dict = Depends(get_current_user)):
    """List all available layers in the GDB"""
    import pyogrio
    
    # Only staff can access
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    if not GDB_PATH.exists():
        raise HTTPException(status_code=404, detail="Base de datos geográfica no disponible")
    
    try:
        layers = pyogrio.list_layers(str(GDB_PATH))
        return {
            "capas": [{"nombre": layer[0], "tipo_geometria": layer[1]} for layer in layers],
            "total": len(layers)
        }
    except Exception as e:
        logger.error(f"Error listing GDB layers: {e}")
        raise HTTPException(status_code=500, detail=f"Error listando capas: {str(e)}")


# ===== NOTIFICACIONES Y ALERTAS GDB =====

@api_router.get("/notificaciones")
async def get_notificaciones(
    leidas: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las notificaciones del usuario actual"""
    query = {"usuario_id": current_user['id']}
    if leidas is not None:
        query["leida"] = leidas
    
    notificaciones = await db.notificaciones.find(query, {"_id": 0}).sort("fecha", -1).limit(50).to_list(50)
    no_leidas = await db.notificaciones.count_documents({"usuario_id": current_user['id'], "leida": False})
    
    return {
        "notificaciones": notificaciones,
        "no_leidas": no_leidas
    }

@api_router.patch("/notificaciones/{notificacion_id}/leer")
async def marcar_notificacion_leida(
    notificacion_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marca una notificación como leída"""
    result = await db.notificaciones.update_one(
        {"id": notificacion_id, "usuario_id": current_user['id']},
        {"$set": {"leida": True, "fecha_lectura": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    
    return {"message": "Notificación marcada como leída"}

@api_router.post("/notificaciones/marcar-todas-leidas")
async def marcar_todas_leidas(current_user: dict = Depends(get_current_user)):
    """Marca todas las notificaciones del usuario como leídas"""
    result = await db.notificaciones.update_many(
        {"usuario_id": current_user['id'], "leida": False},
        {"$set": {"leida": True, "fecha_lectura": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"{result.modified_count} notificaciones marcadas como leídas"}

async def crear_notificacion(usuario_id: str, titulo: str, mensaje: str, tipo: str = "info", enviar_email: bool = False):
    """Crea una notificación para un usuario y opcionalmente envía email"""
    notificacion = {
        "id": str(uuid.uuid4()),
        "usuario_id": usuario_id,
        "titulo": titulo,
        "mensaje": mensaje,
        "tipo": tipo,  # info, warning, success, error
        "leida": False,
        "fecha": datetime.now(timezone.utc).isoformat()
    }
    await db.notificaciones.insert_one(notificacion)
    
    # Enviar email si está habilitado
    if enviar_email:
        user = await db.users.find_one({"id": usuario_id}, {"_id": 0, "email": 1, "full_name": 1})
        if user and user.get('email'):
            try:
                await send_notification_email(user['email'], user.get('full_name', ''), titulo, mensaje)
            except Exception as e:
                logger.error(f"Error enviando email de notificación: {e}")
    
    return notificacion

async def send_notification_email(to_email: str, to_name: str, subject: str, message: str):
    """Envía un email de notificación"""
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        msg['Subject'] = f"[Asomunicipios] {subject}"
        
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #ddd; border-radius: 8px; padding: 20px;">
                <div style="text-align: center; margin-bottom: 20px;">
                    <h2 style="color: #059669;">Asomunicipios</h2>
                </div>
                <p>Hola {to_name},</p>
                <div style="background-color: #f0fdf4; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin-top: 0; color: #166534;">{subject}</h3>
                    <p>{message}</p>
                </div>
                <p style="color: #666; font-size: 12px;">Este es un mensaje automático del sistema de gestión catastral.</p>
            </div>
        </body>
        </html>
        """
        msg.attach(MIMEText(html_body, 'html'))
        
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
            
        logger.info(f"Email de notificación enviado a {to_email}")
    except Exception as e:
        logger.error(f"Error enviando email: {e}")
        raise

@api_router.post("/gdb/enviar-alertas-mensuales")
async def enviar_alertas_mensuales_gdb(current_user: dict = Depends(get_current_user)):
    """Envía alertas mensuales a los gestores con permiso GDB (ejecutar el día 1 de cada mes)"""
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden enviar alertas")
    
    # Buscar gestores con permiso GDB
    gestores_gdb = await db.users.find(
        {"puede_actualizar_gdb": True, "role": UserRole.GESTOR},
        {"_id": 0}
    ).to_list(100)
    
    alertas_enviadas = 0
    mes_actual = datetime.now().strftime("%B %Y")
    
    for gestor in gestores_gdb:
        titulo = "Recordatorio: Cargar Base Gráfica Mensual"
        mensaje = f"Es momento de cargar la base gráfica (GDB) correspondiente al mes de {mes_actual}. Por favor, acceda a Gestión de Predios > Base Gráfica para realizar la carga."
        
        await crear_notificacion(
            usuario_id=gestor['id'],
            titulo=titulo,
            mensaje=mensaje,
            tipo="warning",
            enviar_email=True
        )
        alertas_enviadas += 1
    
    return {
        "message": f"Alertas enviadas a {alertas_enviadas} gestores",
        "gestores_notificados": [g['full_name'] for g in gestores_gdb]
    }

@api_router.get("/gdb/verificar-alerta-mensual")
async def verificar_alerta_mensual(current_user: dict = Depends(get_current_user)):
    """Verifica si es día 1 del mes y si se debe mostrar alerta de carga GDB"""
    hoy = datetime.now()
    es_dia_1 = hoy.day == 1
    
    # Verificar si el usuario tiene permiso GDB
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    tiene_permiso_gdb = user_db.get('puede_actualizar_gdb', False) if user_db else False
    
    # Verificar si ya cargó este mes
    mes_actual = hoy.strftime("%Y-%m")
    carga_este_mes = await db.gdb_cargas.find_one({
        "mes": mes_actual,
        "uploaded_by": current_user['id']
    })
    
    mostrar_alerta = es_dia_1 and tiene_permiso_gdb and not carga_este_mes
    
    return {
        "es_dia_1": es_dia_1,
        "tiene_permiso_gdb": tiene_permiso_gdb,
        "ya_cargo_este_mes": carga_este_mes is not None,
        "mostrar_alerta": mostrar_alerta,
        "mes_actual": mes_actual
    }

@api_router.post("/gdb/upload")
async def upload_gdb_file(
    files: List[UploadFile] = File(...),
    municipio: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload GDB files (ZIP or multiple files from a GDB folder). Only authorized gestors can do this."""
    import zipfile
    import shutil
    import geopandas as gpd
    
    # Check if user is an authorized gestor
    user_db = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    
    if not user_db:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Only gestors with puede_actualizar_gdb permission or coordinador/admin can upload
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        if current_user['role'] != UserRole.GESTOR or not user_db.get('puede_actualizar_gdb', False):
            raise HTTPException(
                status_code=403, 
                detail="No tiene permiso para actualizar la base gráfica. Contacte al coordinador."
            )
    
    try:
        gdb_data_dir = Path("/app/gdb_data")
        gdb_data_dir.mkdir(exist_ok=True)
        
        gdb_found = None
        is_zip = len(files) == 1 and files[0].filename.endswith('.zip')
        
        if is_zip:
            # Proceso ZIP tradicional
            file = files[0]
            temp_zip = UPLOAD_DIR / f"temp_gdb_{uuid.uuid4()}.zip"
            with open(temp_zip, 'wb') as f:
                content = await file.read()
                f.write(content)
            
            with zipfile.ZipFile(temp_zip, 'r') as zip_ref:
                zip_ref.extractall(gdb_data_dir)
            
            temp_zip.unlink()
            
            # Buscar carpeta .gdb
            for item in gdb_data_dir.iterdir():
                if item.suffix == '.gdb' and item.is_dir():
                    gdb_found = item
                    break
            
            if not gdb_found:
                for item in gdb_data_dir.iterdir():
                    if item.is_dir():
                        for subitem in item.iterdir():
                            if subitem.suffix == '.gdb' and subitem.is_dir():
                                gdb_found = subitem
                                break
        else:
            # Proceso para archivos de carpeta GDB (múltiples archivos)
            # Determinar el nombre de la carpeta .gdb desde los archivos
            gdb_folder_name = None
            for file in files:
                # Los archivos vienen con path relativo como "54003.gdb/archivo.ext"
                parts = file.filename.split('/')
                if len(parts) > 0:
                    for part in parts:
                        if part.endswith('.gdb'):
                            gdb_folder_name = part
                            break
                if gdb_folder_name:
                    break
            
            if not gdb_folder_name:
                # Intentar extraer del nombre del archivo
                for file in files:
                    if '.gdb' in file.filename:
                        idx = file.filename.find('.gdb')
                        start = file.filename.rfind('/', 0, idx)
                        gdb_folder_name = file.filename[start+1:idx+4]
                        break
            
            if not gdb_folder_name:
                gdb_folder_name = f"{municipio or 'uploaded'}.gdb"
            
            gdb_found = gdb_data_dir / gdb_folder_name
            gdb_found.mkdir(exist_ok=True)
            
            # Guardar todos los archivos
            for file in files:
                # Extraer solo el nombre del archivo (sin la ruta de carpeta GDB)
                filename = file.filename
                if '/' in filename:
                    filename = filename.split('/')[-1]
                elif '\\' in filename:
                    filename = filename.split('\\')[-1]
                
                file_path = gdb_found / filename
                content = await file.read()
                with open(file_path, 'wb') as f:
                    f.write(content)
        
        if not gdb_found or not gdb_found.exists():
            raise HTTPException(status_code=400, detail="No se pudo crear/encontrar el archivo .gdb")
        
        # Determinar código de municipio desde el nombre del GDB
        gdb_name = gdb_found.stem  # ej: "54003"
        
        # Mapeo de códigos a nombres de municipio
        CODIGO_TO_MUNICIPIO = {
            '54003': 'Ábrego',
            '54109': 'Bucarasica', 
            '54128': 'Cáchira',
            '54206': 'Convención',
            '54245': 'El Carmen',
            '54250': 'El Tarra',
            '54344': 'Hacarí',
            '54398': 'La Playa',
            '54670': 'San Calixto',
            '54720': 'Sardinata',
            '54800': 'Teorama',
            '20614': 'Río de Oro',
        }
        
        municipio_nombre = municipio or CODIGO_TO_MUNICIPIO.get(gdb_name, gdb_name)
        
        # Leer capas del GDB para obtener estadísticas y relacionar con predios
        stats = {"rurales": 0, "urbanos": 0, "relacionados": 0}
        codigos_gdb = set()
        
        try:
            # Intentar diferentes nombres de capas
            for rural_layer in ['R_TERRENO_1', 'R_TERRENO', 'R_Terreno']:
                try:
                    gdf_rural = gpd.read_file(str(gdb_found), layer=rural_layer)
                    stats["rurales"] = len(gdf_rural)
                    # Extraer códigos prediales
                    for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO']:
                        if col in gdf_rural.columns:
                            codigos_gdb.update(gdf_rural[col].dropna().astype(str).tolist())
                            break
                    break
                except:
                    continue
            
            for urban_layer in ['U_TERRENO_1', 'U_TERRENO', 'U_Terreno']:
                try:
                    gdf_urban = gpd.read_file(str(gdb_found), layer=urban_layer)
                    stats["urbanos"] = len(gdf_urban)
                    for col in ['CODIGO', 'codigo', 'CODIGO_PREDIAL', 'codigo_predial', 'COD_PREDIO']:
                        if col in gdf_urban.columns:
                            codigos_gdb.update(gdf_urban[col].dropna().astype(str).tolist())
                            break
                    break
                except:
                    continue
        except Exception as e:
            logger.warning(f"Error leyendo capas GDB: {e}")
        
        # Relacionar con predios existentes
        if codigos_gdb:
            # Actualizar predios con referencia a GDB
            result = await db.predios.update_many(
                {"codigo_predial_nacional": {"$in": list(codigos_gdb)}},
                {"$set": {"tiene_geometria": True, "gdb_source": gdb_name, "gdb_updated": datetime.now(timezone.utc).isoformat()}}
            )
            stats["relacionados"] = result.modified_count
        
        # Registrar carga mensual
        mes_actual = datetime.now().strftime("%Y-%m")
        await db.gdb_cargas.update_one(
            {"mes": mes_actual, "municipio": municipio_nombre},
            {"$set": {
                "id": str(uuid.uuid4()),
                "mes": mes_actual,
                "municipio": municipio_nombre,
                "gdb_file": gdb_name,
                "uploaded_by": current_user['id'],
                "uploaded_by_name": current_user['full_name'],
                "fecha": datetime.now(timezone.utc).isoformat(),
                "predios_rurales": stats["rurales"],
                "predios_urbanos": stats["urbanos"],
                "predios_relacionados": stats["relacionados"]
            }},
            upsert=True
        )
        
        # Notificar a coordinadores que se completó la carga
        coordinadores = await db.users.find(
            {"role": {"$in": [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]}},
            {"_id": 0, "id": 1, "full_name": 1}
        ).to_list(20)
        
        for coord in coordinadores:
            await crear_notificacion(
                usuario_id=coord['id'],
                titulo=f"Base Gráfica Cargada - {municipio_nombre}",
                mensaje=f"{current_user['full_name']} ha cargado la base gráfica de {municipio_nombre} para {mes_actual}. Total geometrías: {stats['rurales'] + stats['urbanos']}, predios relacionados: {stats['relacionados']}",
                tipo="success",
                enviar_email=True
            )
        
        return {
            "message": f"Base gráfica de {municipio_nombre} actualizada exitosamente",
            "municipio": municipio_nombre,
            "gdb_file": gdb_name,
            "predios_rurales": stats["rurales"],
            "predios_urbanos": stats["urbanos"],
            "total_geometrias": stats["rurales"] + stats["urbanos"],
            "predios_relacionados": stats["relacionados"]
        }
        
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="El archivo ZIP no es válido")
    except Exception as e:
        logger.error(f"Error uploading GDB: {e}")
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")


@api_router.get("/gdb/cargas-mensuales")
async def get_cargas_mensuales_gdb(
    mes: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene las cargas de GDB del mes actual o especificado"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    mes_consulta = mes or datetime.now().strftime("%Y-%m")
    
    cargas = await db.gdb_cargas.find(
        {"mes": mes_consulta},
        {"_id": 0}
    ).to_list(50)
    
    return {
        "mes": mes_consulta,
        "total_cargas": len(cargas),
        "cargas": cargas
    }


@api_router.get("/gdb/predios-con-geometria")
async def get_predios_con_geometria(
    municipio: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtiene estadísticas de predios que tienen geometría GDB asociada"""
    if current_user['role'] == UserRole.CIUDADANO:
        raise HTTPException(status_code=403, detail="No tiene permiso")
    
    query = {"tiene_geometria": True}
    if municipio:
        query["municipio"] = municipio
    
    total_con_geometria = await db.predios.count_documents(query)
    total_predios = await db.predios.count_documents({"municipio": municipio} if municipio else {})
    
    # Por municipio
    pipeline = [
        {"$match": {"tiene_geometria": True}},
        {"$group": {"_id": "$municipio", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    por_municipio = await db.predios.aggregate(pipeline).to_list(50)
    
    return {
        "total_con_geometria": total_con_geometria,
        "total_predios": total_predios,
        "porcentaje": round((total_con_geometria / total_predios * 100) if total_predios > 0 else 0, 2),
        "por_municipio": [{"municipio": r["_id"], "count": r["count"]} for r in por_municipio]
    }


@api_router.patch("/users/{user_id}/gdb-permission")
async def update_user_gdb_permission(
    user_id: str,
    puede_actualizar: bool,
    current_user: dict = Depends(get_current_user)
):
    """Allow coordinador to grant/revoke GDB update permission to a gestor"""
    # Only coordinador or admin can grant this permission
    if current_user['role'] not in [UserRole.COORDINADOR, UserRole.ADMINISTRADOR]:
        raise HTTPException(status_code=403, detail="Solo coordinadores pueden asignar este permiso")
    
    # Find user
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Only gestors can have this permission
    if user['role'] != UserRole.GESTOR:
        raise HTTPException(status_code=400, detail="Este permiso solo aplica para gestores")
    
    # Update permission
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"puede_actualizar_gdb": puede_actualizar}}
    )
    
    return {
        "message": f"Permiso {'otorgado' if puede_actualizar else 'revocado'} exitosamente",
        "user_id": user_id,
        "puede_actualizar_gdb": puede_actualizar
    }


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
